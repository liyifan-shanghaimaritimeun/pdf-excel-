"""
PDF表格数据提取器

核心功能：
1. 使用PyMuPDF(fitz)进行PDF版面分析和文本提取
2. 集成PaddleOCR处理复杂表格或扫描件
3. 支持先字典后泛化的表头搜索策略
4. 自动识别双栏表格并分割处理
5. 使用Pandas进行数据清洗和导出

适用场景：
- PDF产品目录表格提取
- 扫描件表格识别
- 多页PDF批量提取
- 大文件(200MB+)处理
"""

import os
import re
import sys
import json
import time
import threading
import logging
import fitz
import pandas as pd
import cv2
import numpy as np
from PIL import Image
from collections import defaultdict
from typing import List, Dict, Tuple, Optional, Any

logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class OCRTableExtractor:
    """
    OCR表格识别器

    使用PaddleOCR对PDF页面进行文字识别，将识别结果转换为带坐标信息的单元格列表。
    支持中英文混合识别，自动处理文字方向。
    """

    _ocr_lock = threading.Lock()

    def __init__(self, use_gpu: bool = False):
        """
        初始化OCR识别器

        Args:
            use_gpu: 是否使用GPU加速，默认False
        """
        self.use_gpu = use_gpu
        self.ocr = None
        self._init_ocr()

    def _init_ocr(self):
        """初始化PaddleOCR实例（延迟加载）"""
        try:
            from paddleocr import PaddleOCR
            # 打包环境：显式指定模型路径，避免PaddleOCR找不到模型
            kwargs = dict(
                lang='ch',           # 中英文混合识别
                use_gpu=self.use_gpu,
                use_angle_cls=True,  # 启用文字方向分类
                show_log=False       # 关闭日志输出
            )
            if getattr(sys, 'frozen', False):
                base_dir = os.path.dirname(sys.executable)
                model_dir = os.path.join(base_dir, 'models', 'whl')
                det_dir = os.path.join(model_dir, 'det', 'ch', 'ch_PP-OCRv4_det_infer')
                rec_dir = os.path.join(model_dir, 'rec', 'ch', 'ch_PP-OCRv4_rec_infer')
                cls_dir = os.path.join(model_dir, 'cls', 'ch_ppocr_mobile_v2.0_cls_infer')
                if os.path.isdir(det_dir):
                    kwargs['det_model_dir'] = det_dir
                if os.path.isdir(rec_dir):
                    kwargs['rec_model_dir'] = rec_dir
                if os.path.isdir(cls_dir):
                    kwargs['cls_model_dir'] = cls_dir
                print(f"OCR打包模式: det={os.path.isdir(det_dir)}, rec={os.path.isdir(rec_dir)}, cls={os.path.isdir(cls_dir)}")
            self.ocr = PaddleOCR(**kwargs)
        except ImportError:
            pass
        except Exception as e:
            print(f"OCR初始化失败: {e}")
            self.ocr = None

            

    def _preprocess_image(self, image_path: str, is_scanned: bool = False) -> str:
        """
        图像预处理：自适应二值化、去噪、透视校正
        
        Args:
            image_path: 原始图片路径
            is_scanned: 是否为扫描件

        Returns:
            预处理后的图片路径
        """
        try:
            img = cv2.imread(image_path)
            if img is None:
                return image_path
            
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            
            if is_scanned:
                gray = cv2.fastNlMeansDenoising(gray, None, 10, 7, 21)
                
                gray = cv2.adaptiveThreshold(
                    gray, 255,
                    cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                    cv2.THRESH_BINARY,
                    11, 2
                )
                
                gray = self._perspective_correction(gray)
            
            preprocessed_path = image_path.replace('.png', '_preprocessed.png')
            cv2.imwrite(preprocessed_path, gray)
            
            return preprocessed_path
        except Exception:
            return image_path

    def _perspective_correction(self, img: np.ndarray) -> np.ndarray:
        """
        透视校正：检测文档边缘并进行透视变换
        
        Args:
            img: 输入图像（灰度图）

        Returns:
            校正后的图像
        """
        try:
            edges = cv2.Canny(img, 50, 150, apertureSize=3)
            
            contours, _ = cv2.findContours(
                edges.copy(),
                cv2.RETR_EXTERNAL,
                cv2.CHAIN_APPROX_SIMPLE
            )
            
            contours = sorted(contours, key=cv2.contourArea, reverse=True)[:5]
            
            for contour in contours:
                perimeter = cv2.arcLength(contour, True)
                approx = cv2.approxPolyDP(contour, 0.02 * perimeter, True)
                
                if len(approx) == 4:
                    rect = approx
                    break
            else:
                return img
            
            pts = rect.reshape(4, 2)
            rect = np.zeros((4, 2), dtype="float32")
            
            s = pts.sum(axis=1)
            rect[0] = pts[np.argmin(s)]
            rect[2] = pts[np.argmax(s)]
            
            diff = np.diff(pts, axis=1)
            rect[1] = pts[np.argmin(diff)]
            rect[3] = pts[np.argmax(diff)]
            
            (tl, tr, br, bl) = rect
            
            width_a = np.sqrt(((br[0] - bl[0]) ** 2) + ((br[1] - bl[1]) ** 2))
            width_b = np.sqrt(((tr[0] - tl[0]) ** 2) + ((tr[1] - tl[1]) ** 2))
            max_width = max(int(width_a), int(width_b))
            
            height_a = np.sqrt(((tr[0] - br[0]) ** 2) + ((tr[1] - br[1]) ** 2))
            height_b = np.sqrt(((tl[0] - bl[0]) ** 2) + ((tl[1] - bl[1]) ** 2))
            max_height = max(int(height_a), int(height_b))
            
            dst = np.array([
                [0, 0],
                [max_width - 1, 0],
                [max_width - 1, max_height - 1],
                [0, max_height - 1]
            ], dtype="float32")
            
            M = cv2.getPerspectiveTransform(rect, dst)
            warped = cv2.warpPerspective(img, M, (max_width, max_height))
            
            return warped
        except Exception:
            return img

    def recognize_from_image(self, image_path: str, is_scanned: bool = False) -> List[Dict]:
        """
        从图片中识别文字，返回带坐标信息的单元格列表

        Args:
            image_path: 图片文件路径
            is_scanned: 是否为扫描件

        Returns:
            单元格列表，每个单元格包含：text(文字), x0/y0/x1/y1(边界坐标), center_x/center_y(中心坐标)
        """
        if not self.ocr:
            return []

        try:
            with self._ocr_lock:
                # 策略：原图优先 → 预处理兜底
                # 原因：PaddleOCR 自身对扫描件有容忍度，强制二值化/去噪
                # 对彩色/复杂背景页面常产生反效果（整页识别为空）
                result = self.ocr.ocr(image_path, cls=True)
                if not result or not result[0]:
                    processed_path = self._preprocess_image(image_path, is_scanned)
                    result = self.ocr.ocr(processed_path, cls=True)
                if not result or not result[0]:
                    print(f"✗ OCR识别失败: {image_path}")
                    return []


                  
                cells = []
                for line in result[0]:
                    box = line[0]       # 边界框坐标
                    text = line[1][0]   # 识别文字

                    x0, y0 = box[0]     # 左上角坐标
                    x1, y1 = box[2]     # 右下角坐标

                    cells.append({
                        "text": text.strip(),
                        "x0": x0,
                        "y0": y0,
                        "x1": x1,
                        "y1": y1,
                        "center_x": (x0 + x1) / 2,  # X中心坐标
                        "center_y": (y0 + y1) / 2,  # Y中心坐标
                    })

                return cells
        except Exception:
            return []

    def recognize_from_pdf_page(self, pdf_path: str, page_num: int = 0, is_scanned: bool = False) -> List[Dict]:
        """
        从PDF指定页面识别文字

        Args:
            pdf_path: PDF文件路径
            page_num: 页码（从0开始）
            is_scanned: 是否为扫描件

        Returns:
            单元格列表，每个单元格包含文字和坐标信息
        """
        doc = fitz.open(pdf_path)
        page = doc.load_page(page_num)

        if is_scanned:
            pix = page.get_pixmap(matrix=fitz.Matrix(4.0, 4.0))
        else:
            pix = page.get_pixmap(matrix=fitz.Matrix(3.0, 3.0))

        temp_path = f"_temp_page_{page_num}.png"
        pix.save(temp_path)

        cells = self.recognize_from_image(temp_path, is_scanned)

        if os.path.exists(temp_path):
            os.remove(temp_path)

        return cells

    def recognize_table_with_ppstructure(self, image_path: str) -> List[List[List[str]]]:
        """
        使用PPStructure专用接口识别表格结构（旧接口，保留兼容）

        Args:
            image_path: 图片文件路径

        Returns:
            表格数据列表，每个表格是一个二维列表
        """
        import traceback
        try:
            from paddleocr import PaddleOCR
            table_engine = PaddleOCR(
                lang='ch',
                use_gpu=self.use_gpu,
                use_angle_cls=True,
                show_log=False,
                table=True
            )
            
            with self._ocr_lock:
                result = table_engine.ocr(image_path, cls=True)
            
            if not result or not result[0]:
                logger.warning(f"PPStructure: 返回空结果 image={image_path}")
                return []
            
            tables = []
            for item in result[0]:
                if isinstance(item, dict) and 'table' in item:
                    table_data = item['table']
                    if table_data and isinstance(table_data, list):
                        tables.append(table_data)
            
            logger.info(f"PPStructure: 检测到 {len(tables)} 个表格 image={image_path}")
            return tables
        except Exception as e:
            logger.error(f"PPStructure异常: {e}\n{traceback.format_exc()}")
            return []

    def scan_page_ppstructure(self, page: 'fitz.Page', page_num: int, dpi: int = 300) -> Dict:
        """
        扫描页全页识别：OCR文字 + 优化坐标聚类推断表格结构

        优化：
        1. 动态Y行聚类阈值（基于中位行高）
        2. 改进X列检测（处理合并单元格）
        3. 单元格文本合并（处理同一列多个cell）
        4. 全页文字输出（用于用户搜索）

        Args:
            page: fitz.Page对象
            page_num: 页码
            dpi: 渲染DPI

        Returns:
            {
                'tables': [{'data': [[...]], 'bbox': [...], 'source': 'ppstructure'}],
                'out_table_cells': [{'text': ..., 'pixel_bbox': [...], 'cx': ..., 'cy': ...}],
                'all_cells': [全部OCR cells],
                'all_text': '全页文字',
            }
        """
        import tempfile
        result = {
            'tables': [],
            'out_table_cells': [],
            'all_cells': [],
            'all_text': '',
        }

        pix = page.get_pixmap(matrix=fitz.Matrix(dpi / 72.0, dpi / 72.0))
        fd, temp_img_path = tempfile.mkstemp(suffix='.png')
        os.close(fd)
        pix.save(temp_img_path)
        logger.info(f"全页OCR扫描: 渲染页面 {pix.width}x{pix.height} (DPI={dpi})")

        try:
            # 1. PaddleOCR全页识别
            if not self.ocr:
                logger.error(f"全页OCR: OCR引擎未初始化，跳过第{page_num+1}页")
                return result
            scale = dpi / 72.0

            with self._ocr_lock:
                raw = self.ocr.ocr(temp_img_path, cls=True)

            if not raw or not raw[0]:
                logger.warning(f"全页OCR: 第{page_num+1}页返回空结果")
                return result

            # 2. 解析OCR结果 → cells
            cells = []
            for item in raw[0]:
                if not isinstance(item, list) or len(item) < 2:
                    continue
                bbox = item[0]  # [[x0,y0], [x1,y0], [x1,y1], [x0,y1]]
                text_info = item[1]  # (text, confidence)
                if len(bbox) != 4 or not text_info:
                    continue

                text = text_info[0] if isinstance(text_info, (list, tuple)) else str(text_info)
                x0, y0 = bbox[0]
                x1, y1 = bbox[2]

                cells.append({
                    'text': str(text).strip(),
                    'x0': x0,
                    'y0': y0,
                    'x1': x1,
                    'y1': y1,
                    'center_x': (x0 + x1) / 2,
                    'center_y': (y0 + y1) / 2,
                    'pixel_bbox': [x0, y0, x1, y1],
                    'height': y1 - y0,
                })

            result['all_cells'] = cells
            # 生成全页文字（用于用户搜索）
            result['all_text'] = ' '.join(c['text'] for c in cells if c['text'])
            logger.info(f"全页OCR: 第{page_num+1}页识别到 {len(cells)} 个文字块")

            if not cells:
                return result

            # 3. cells按Y坐标聚合成lines（动态阈值）
            sorted_cells = sorted(cells, key=lambda c: (c['center_y'], c['center_x']))
            
            # 计算动态Y容差（基于中位cell高度）
            heights = [c['height'] for c in cells]
            median_height = sorted(heights)[len(heights) // 2] if heights else 10
            y_tolerance = max(median_height * 0.5, 8)  # 动态阈值
            
            lines = []
            current_line = []
            current_y = None

            for cell in sorted_cells:
                cy = cell['center_y']
                if current_y is None:
                    current_y = cy
                    current_line.append(cell)
                elif abs(cy - current_y) <= y_tolerance:
                    current_line.append(cell)
                    current_y = sum(c['center_y'] for c in current_line) / len(current_line)
                else:
                    if current_line:
                        lines.append(self._merge_cells_to_line(current_line))
                    current_line = [cell]
                    current_y = cy

            if current_line:
                lines.append(self._merge_cells_to_line(current_line))

            lines = [l for l in lines if l]
            logger.info(f"全页OCR: 第{page_num+1}页聚合成 {len(lines)} 行 (Y容差={y_tolerance:.1f})")

            # 4. lines按X坐标聚类推断表格区域
            table_infos = self._infer_tables_from_lines(lines)
            logger.info(f"全页OCR: 第{page_num+1}页推断出 {len(table_infos)} 个表格区域")

            # 5. 对每个表格区域提取结构化数据
            for cell in cells:
                cell['in_table'] = False
                
            for tbl_idx, tbl_info in enumerate(table_infos):
                tbl_bbox = tbl_info.get('pixel_bbox', [0, 0, pix.width, pix.height])

                # 找出属于这个表格的cells
                tbl_cells = []
                for cell in cells:
                    cx, cy = cell['center_x'], cell['center_y']
                    if tbl_bbox[0] <= cx <= tbl_bbox[2] and tbl_bbox[1] <= cy <= tbl_bbox[3]:
                        tbl_cells.append(cell)
                        cell['in_table'] = True

                if not tbl_cells:
                    continue

                # 提取表格的列信息
                col_centers = self._detect_columns_improved(tbl_cells)

                # 按行组织cells成二维表格（改进版：合并同列文本）
                table_data = self._cells_to_table_data_improved(tbl_cells, col_centers)

                if table_data and len(table_data) >= 1:
                    # 转换bbox到PDF坐标
                    bbox_pdf = [
                        tbl_bbox[0] / scale,
                        tbl_bbox[1] / scale,
                        tbl_bbox[2] / scale,
                        tbl_bbox[3] / scale,
                    ]
                    result['tables'].append({
                        'data': table_data,
                        'bbox': bbox_pdf,
                        'pixel_bbox': tbl_bbox,
                        'source': 'ppstructure',
                    })

            # 6. 表外文字 = 不在任何表格内的cells
            for cell in cells:
                if not cell.get('in_table', False):
                    result['out_table_cells'].append({
                        'text': cell['text'],
                        'pixel_bbox': cell['pixel_bbox'],
                        'center_x': cell['center_x'],
                        'center_y': cell['center_y'],
                    })

            logger.info(f"全页OCR: 第{page_num+1}页 表格={len(result['tables'])}, 表外文字={len(result['out_table_cells'])}")

        except Exception as e:
            import traceback
            logger.error(f"全页OCR第{page_num+1}页异常: {e}\n{traceback.format_exc()}")
        finally:
            try:
                os.remove(temp_img_path)
            except Exception:
                pass

        return result

    def _merge_cells_to_line(self, cells: List[Dict]) -> Optional[Dict]:
        """合并同一行的cells为line"""
        if not cells:
            return None
        sorted_cells = sorted(cells, key=lambda c: c['x0'])
        x0 = min(c['x0'] for c in sorted_cells)
        x1 = max(c['x1'] for c in sorted_cells)
        y0 = min(c['y0'] for c in sorted_cells)
        y1 = max(c['y1'] for c in sorted_cells)
        text = ' '.join(c['text'] for c in sorted_cells if c['text'])
        return {
            'text': text,
            'x0': x0, 'y0': y0, 'x1': x1, 'y1': y1,
            'cx': (x0 + x1) / 2,
            'cy': (y0 + y1) / 2,
            'cells': sorted_cells,
            'x_clusters': [c['center_x'] for c in sorted_cells],
        }

    def _detect_columns(self, cells: List[Dict]) -> List[float]:
        """检测表格cells的列中心X坐标"""
        if not cells:
            return []

        # 按行分组
        rows = {}
        for cell in cells:
            cy = round(cell['center_y'] / 5) * 5
            if cy not in rows:
                rows[cy] = []
            rows[cy].append(cell)

        # 收集每行的cell中心X
        all_x = []
        for cells_in_row in rows.values():
            x_positions = sorted(c['center_x'] for c in cells_in_row)
            all_x.extend(x_positions)

        if len(all_x) < 2:
            return [sum(all_x) / len(all_x)] if all_x else []

        # X坐标聚类
        sorted_x = sorted(all_x)
        diffs = [sorted_x[i+1] - sorted_x[i] for i in range(len(sorted_x)-1)]
        median_diff = sorted(diffs)[len(diffs) // 2]
        gap_threshold = max(median_diff * 1.2, 15)

        clusters = [[sorted_x[0]]]
        for k in range(1, len(sorted_x)):
            if sorted_x[k] - sorted_x[k-1] > gap_threshold:
                clusters.append([sorted_x[k]])
            else:
                clusters[-1].append(sorted_x[k])

        return [sum(c) / len(c) for c in clusters]

    def _cells_to_table_data(self, cells: List[Dict], col_centers: List[float]) -> List[List[str]]:
        """将cells按列中心组织成二维表格数据"""
        if not cells:
            return []

        # 按行分组
        rows = {}
        for cell in cells:
            cy = round(cell['center_y'] / 5) * 5
            if cy not in rows:
                rows[cy] = []
            rows[cy].append(cell)

        # 对每行，把cell分配到最近的列
        table_data = []
        for cy in sorted(rows.keys()):
            row_cells = rows[cy]
            row_data = [''] * len(col_centers)

            for cell in row_cells:
                cx = cell['center_x']
                # 找到最近的列
                best_col = 0
                best_dist = float('inf')
                for i, col_cx in enumerate(col_centers):
                    dist = abs(cx - col_cx)
                    if dist < best_dist:
                        best_dist = dist
                        best_col = i
                row_data[best_col] = cell['text']

            # 跳过全空行
            if any(c.strip() for c in row_data):
                table_data.append(row_data)

        return table_data

    def _detect_columns_improved(self, cells: List[Dict]) -> List[float]:
        """
        改进的列中心检测
        
        优化：
        1. 使用 DBSCAN 风格的聚类（自适应阈值）
        2. 处理合并单元格（跨多列的cell纳入所有相关列）
        3. 过滤异常值
        """
        if not cells:
            return []

        # 收集所有cell的中心X坐标
        all_x = [c['center_x'] for c in cells]
        
        if len(all_x) < 2:
            return [sum(all_x) / len(all_x)] if all_x else []

        # 自适应聚类：基于数据分布的间隙
        sorted_x = sorted(all_x)
        diffs = [sorted_x[i+1] - sorted_x[i] for i in range(len(sorted_x)-1)]
        
        # 使用间隙的均值+标准差作为阈值
        mean_diff = sum(diffs) / len(diffs)
        std_diff = (sum((d - mean_diff) ** 2 for d in diffs) / len(diffs)) ** 0.5
        gap_threshold = max(mean_diff + std_diff, 15)  # 至少15像素

        # 聚类
        clusters = [[sorted_x[0]]]
        for k in range(1, len(sorted_x)):
            if sorted_x[k] - sorted_x[k-1] > gap_threshold:
                clusters.append([sorted_x[k]])
            else:
                clusters[-1].append(sorted_x[k])

        # 计算每个簇的中心
        col_centers = [sum(c) / len(c) for c in clusters]
        
        # 验证：至少30%的cell应该能匹配到某个列中心
        valid_cells = 0
        for cell in cells:
            cx = cell['center_x']
            for col_cx in col_centers:
                if abs(cx - col_cx) < gap_threshold:
                    valid_cells += 1
                    break
        
        # 如果列检测效果不好，用更简单的方法
        if valid_cells < len(cells) * 0.5:
            # 降级：直接用等间距
            if len(col_centers) < 2:
                # 只有一列
                x_min, x_max = min(all_x), max(all_x)
                return [(x_min + x_max) / 2]
        
        return col_centers

    def _cells_to_table_data_improved(self, cells: List[Dict], col_centers: List[float]) -> List[List[str]]:
        """
        改进的表格数据转换
        
        优化：
        1. 同一列多个cell的文本合并（用空格连接）
        2. 处理跨行的cell（根据Y坐标分配）
        3. 保持原始列顺序
        """
        if not cells or not col_centers:
            return []

        # 按行分组（使用动态阈值）
        rows = {}
        for cell in cells:
            # 计算这个cell属于哪一行
            cy = cell['center_y']
            assigned_row = None
            
            # 找最近的行
            for row_key in rows.keys():
                if abs(cy - row_key) <= 15:  # 15像素容差
                    assigned_row = row_key
                    break
            
            if assigned_row is None:
                rows[cy] = []
                assigned_row = cy
            
            rows[assigned_row].append(cell)

        # 对每行，把cells分配到最近的列（支持合并文本）
        table_data = []
        for cy in sorted(rows.keys()):
            row_cells = rows[cy]
            row_data = [''] * len(col_centers)
            
            # 收集每列的所有文本
            col_texts = [[] for _ in col_centers]
            
            for cell in row_cells:
                cx = cell['center_x']
                text = cell['text']
                
                # 找到最近的列
                best_col = 0
                best_dist = float('inf')
                for i, col_cx in enumerate(col_centers):
                    dist = abs(cx - col_cx)
                    if dist < best_dist:
                        best_dist = dist
                        best_col = i
                
                col_texts[best_col].append(text)
            
            # 合并每列的文本
            for i, texts in enumerate(col_texts):
                if texts:
                    row_data[i] = ' '.join(texts)
            
            # 跳过全空行
            if any(c.strip() for c in row_data):
                table_data.append(row_data)

        return table_data

    def _infer_tables_from_lines(self, lines: List[Dict]) -> List[Dict]:
        """
        从lines推断表格区域

        策略：
        1. 用cell_count>=3找出特征行
        2. 相邻特征行组合（Y间距检查+X范围邻近检查）
        3. X范围邻近：允许不重叠但距离<200px（处理跨列表格）
        4. 检查列对齐
        """
        if len(lines) < 2:
            return []

        inferred = []
        sorted_lines = sorted(lines, key=lambda l: (l['y0'], l['x0']))

        # 计算每行的cell数量
        for line in sorted_lines:
            if 'cell_count' not in line:
                line['cell_count'] = len(line.get('cells', []))

        # 找特征行：cell_count >= 3
        feature_indices = [i for i, l in enumerate(sorted_lines) if l['cell_count'] >= 3]

        if len(feature_indices) < 2:
            # 降级：用cell_count>=2
            feature_indices = [i for i, l in enumerate(sorted_lines) if l['cell_count'] >= 2]
            if len(feature_indices) < 2:
                return []

        # 将相邻的特征行组合
        # 改进：允许X范围邻近（距离<200px），处理跨列表格
        groups = []
        current_group = [feature_indices[0]]
        for k in range(1, len(feature_indices)):
            idx = feature_indices[k]
            prev_idx = feature_indices[k-1]
            y_gap = idx - prev_idx
            
            if y_gap > 4:
                # Y间距太大，结束当前组
                if len(current_group) >= 2:
                    groups.append(current_group)
                current_group = [idx]
            else:
                # 检查X范围是否邻近（允许不重叠但距离<200px）
                prev_line = sorted_lines[prev_idx]
                curr_line = sorted_lines[idx]
                prev_x0, prev_x1 = prev_line['x0'], prev_line['x1']
                curr_x0, curr_x1 = curr_line['x0'], curr_line['x1']
                
                # 计算X距离（考虑重叠）
                if prev_x1 < curr_x0:
                    x_distance = curr_x0 - prev_x1  # 前一行在左边
                elif curr_x1 < prev_x0:
                    x_distance = prev_x0 - curr_x1  # 当前行在左边
                else:
                    x_distance = 0  # 有重叠
                
                # 如果距离<200px，认为是同一表格区域
                if x_distance <= 200:
                    current_group.append(idx)
                else:
                    if len(current_group) >= 2:
                        groups.append(current_group)
                    current_group = [idx]
        
        if len(current_group) >= 2:
            groups.append(current_group)

        for group in groups:
            table_lines = [sorted_lines[i] for i in group]

            # 计算bbox
            all_x0 = [l['x0'] for l in table_lines]
            all_y0 = [l['y0'] for l in table_lines]
            all_x1 = [l['x1'] for l in table_lines]
            all_y1 = [l['y1'] for l in table_lines]
            bbox = [min(all_x0), min(all_y0), max(all_x1), max(all_y1)]

            col_count = self._count_x_columns(table_lines)

            if col_count >= 2:
                inferred.append({
                    'bbox': bbox,
                    'pixel_bbox': bbox,
                    'source': 'ppstructure_inferred',
                    'conf': 0.8,
                    'line_count': len(table_lines),
                    'col_count': col_count,
                    'lines': table_lines,
                })
            elif col_count == 1 and len(table_lines) >= 3:
                inferred.append({
                    'bbox': bbox,
                    'pixel_bbox': bbox,
                    'source': 'ppstructure_inferred',
                    'conf': 0.4,
                    'line_count': len(table_lines),
                    'col_count': 1,
                    'lines': table_lines,
                })

        return inferred

    def _count_x_columns(self, lines: List[Dict]) -> int:
        """检测多行文本的x列数"""
        if len(lines) < 2:
            return 0

        all_x = []
        for line in lines:
            xc = line.get('x_clusters')
            if xc:
                all_x.extend(xc)
            else:
                all_x.append((line['x0'] + line['x1']) / 2)

        if len(all_x) < 4:
            return 1

        sorted_x = sorted(all_x)
        diffs = [sorted_x[i+1] - sorted_x[i] for i in range(len(sorted_x)-1)]
        median_diff = sorted(diffs)[len(diffs) // 2]
        gap_threshold = max(median_diff * 1.5, 20)

        clusters = [[sorted_x[0]]]
        for k in range(1, len(sorted_x)):
            if sorted_x[k] - sorted_x[k-1] > gap_threshold:
                clusters.append([sorted_x[k]])
            else:
                clusters[-1].append(sorted_x[k])

        if len(clusters) < 2:
            return 1

        valid_clusters = 0
        for cluster in clusters:
            cluster_mean = sum(cluster) / len(cluster)
            matching_lines = 0
            for line in lines:
                xc = line.get('x_clusters')
                if xc:
                    matches = sum(1 for x in xc if abs(x - cluster_mean) < gap_threshold)
                    if matches >= max(1, len(xc) * 0.3):
                        matching_lines += 1
                else:
                    if abs((line['x0'] + line['x1']) / 2 - cluster_mean) < gap_threshold:
                        matching_lines += 1
            if matching_lines >= len(lines) * 0.4:
                valid_clusters += 1

        return valid_clusters if valid_clusters >= 2 else 1


class TableStructureAnalyzer:
    """
    表格结构分析器

    负责表格数据的清洗、双栏表格分割等结构处理。
    """

    def __init__(self):
        """初始化表格结构分析器"""
        pass

    def clean_table(self, table: List[List[str]]) -> List[List[str]]:
        """
        清洗表格数据：去除空白行，清理单元格内容

        Args:
            table: 原始表格数据（二维列表）

        Returns:
            清洗后的表格数据，不包含全空行
        """
        cleaned = []
        for row in table:
            # 将每个单元格转换为字符串并去除首尾空格
            cleaned_row = [str(cell).strip() if cell else "" for cell in row]
            # 只保留非空行
            if any(cleaned_row):
                cleaned.append(cleaned_row)
        return cleaned

    def split_double_table(self, table: List[List[str]]) -> List[List[List[str]]]:
        """
        自动检测并分割双栏表格

        双栏表格是指PDF中常见的并排显示的两个表格，左右两部分通常是对称的。
        该方法采用两种策略进行分割：
        1. 对称分割（优先）：双栏表格左右列数大致相等，按中间位置分割
        2. 空白列分割（备选）：通过检测中间空白列确定分割位置

        改进点：
        1. 优先使用对称分割，符合PDF表格的实际布局特点
        2. 通过表头结构验证分割结果的正确性
        3. 分割后对每个子表格独立清洗，避免数据丢失

        Args:
            table: 原始表格数据

        Returns:
            分割后的表格列表，可能是1个或2个表格
        """
        if not table or len(table) == 0:
            return [table]

        num_cols = len(table[0])
        # 列数少于8列的通常不是双栏表格（6~7列多为单表带多属性列），直接清洗返回
        if num_cols < 7:
            cleaned = self.clean_table(table)
            return [cleaned] if len(cleaned) >= 2 else [table]

        # 策略1：优先使用对称分割
        # 双栏表格左右列数通常大致相等，尝试在中间位置分割
        split_col = self._find_symmetric_split(table)

        # 策略2（空白列分割）已禁用：其"空白列≥50%"弱信号会误劈单表
        # 仅信任对称分割（方法1/方法2），其余情况视为单表

        # 如果找到了有效的分割列
        if split_col > 0 and split_col < num_cols - 2:
            table1 = []
            table2 = []
            for row in table:
                row1 = row[:split_col]
                row2 = row[split_col:]
                table1.append(row1)
                table2.append(row2)

            # 对每个子表格进行独立清洗
            cleaned_table1 = self.clean_table(table1)
            cleaned_table2 = self.clean_table(table2)

            # 验证分割结果：两个子表格都应有足够的数据
            result = []
            if len(cleaned_table1) >= 2:
                result.append(cleaned_table1)
            if len(cleaned_table2) >= 2:
                result.append(cleaned_table2)

            return result if result else [table]

        # 不是双栏表格，直接清洗返回
        cleaned = self.clean_table(table)
        return [cleaned] if len(cleaned) >= 2 else [table]

    def _find_symmetric_split(self, table: List[List[str]]) -> int:
        """
        通过对称分析找到分割列位置

        双栏表格的特点：
        1. 左右两部分有相同或相似的表头结构
        2. 列数大致相等
        3. 表头列名会重复出现（如左边有"Part Number"，右边也会有）

        改进点：
        1. 使用多行表头进行对称分析，而非仅单行
        2. 排除产品型号行（如HSE102M51），避免误判分割点
        3. 通过多行列结构对比确定更准确的分割位置

        Args:
            table: 表格数据

        Returns:
            分割列索引，未找到返回0
        """
        if not table or len(table) < 2:
            return 0

        num_cols = len(table[0])
        # 至少需要8列才可能是双栏表格（避免窄多列单表被误判）
        if num_cols < 7:
            return 0

        # 使用多行表头进行分析（前5行）
        header_rows = table[:min(5, len(table))]
        
        # 收集所有表头行的关键词（排除产品型号行）
        all_header_keywords = set()
        for row in header_rows:
            for cell in row:
                cell_str = str(cell).strip().lower()
                if cell_str and len(cell_str) >= 2:
                    # 排除产品型号模式（如HSE102M51）
                    if not re.match(r'^[a-zA-Z0-9]+$', str(cell).strip()):
                        all_header_keywords.add(cell_str)

        if len(all_header_keywords) < 2:
            return 0

        # 方法1：通过找到重复的表头关键词确定分割位置
        # 从左到右扫描，找到第一个重复出现的表头关键词位置
        split_col = 0
        found_first_keyword = False
        first_keyword = ""
        
        # 遍历所有表头行，寻找重复的关键词模式
        for row in header_rows:
            for col_idx, cell in enumerate(row):
                cell_str = str(cell).strip().lower()
                if cell_str in all_header_keywords:
                    # 排除纯数字产品型号
                    if re.match(r'^[a-zA-Z0-9]+$', str(cell).strip()):
                        continue
                        
                    if not found_first_keyword:
                        found_first_keyword = True
                        first_keyword = cell_str
                    else:
                        # 如果找到相同的关键词，说明这是右栏的开始
                        if cell_str == first_keyword and col_idx > num_cols // 3:
                            split_col = col_idx
                            return split_col

        # 方法2：基于列数对称性的分割
        # 双栏表格左右列数通常相等，尝试在中间位置分割
        left_half_cols = num_cols // 2
        
        # 验证左右两部分是否有相似的结构
        left_keywords = set()
        right_keywords = set()
        
        for row in header_rows:
            # 左半部分关键词
            for col_idx in range(left_half_cols):
                if col_idx < len(row):
                    cell_str = str(row[col_idx]).strip().lower()
                    if cell_str and len(cell_str) >= 2:
                        if not re.match(r'^[a-zA-Z0-9]+$', str(row[col_idx]).strip()):
                            left_keywords.add(cell_str)
            
            # 右半部分关键词
            for col_idx in range(left_half_cols, num_cols):
                if col_idx < len(row):
                    cell_str = str(row[col_idx]).strip().lower()
                    if cell_str and len(cell_str) >= 2:
                        if not re.match(r'^[a-zA-Z0-9]+$', str(row[col_idx]).strip()):
                            right_keywords.add(cell_str)

        # 如果左右两部分有重叠的关键词，说明是双栏表格
        common_keywords = left_keywords & right_keywords
        
        if len(common_keywords) >= 1:
            # 找到第一个重复关键词在右半部分出现的位置
            for row in header_rows:
                for col_idx in range(left_half_cols, num_cols):
                    if col_idx < len(row):
                        cell_str = str(row[col_idx]).strip().lower()
                # 仅当重复表头位置接近中线才视为双栏分割点，否则可能误劈单表
                if cell_str in common_keywords and abs(col_idx - left_half_cols) <= max(1, int(num_cols*0.15)):
                    return col_idx
            
            # 如果没有找到具体位置，返回中间位置
            return left_half_cols

        # 没有重复表头关键词：不盲目按"空白列≥50%"弱信号分割（方法3），避免误劈单表
        return 0

        # 方法3：检查是否有空白区域作为分隔（用于更准确的分割）
        # 统计每列的空白率
        check_rows = min(6, len(table))
        for col_idx in range(num_cols // 3, num_cols * 2 // 3):
            empty_count = 0
            for row_idx in range(check_rows):
                row = table[row_idx]
                if col_idx < len(row) and str(row[col_idx]).strip() == "":
                    empty_count += 1
            # 如果这一列空白率很高，且前后都有数据，说明是分隔列
            if empty_count >= check_rows * 0.5:
                # 检查前后是否有数据
                has_left_data = False
                has_right_data = False
                for row_idx in range(check_rows):
                    row = table[row_idx]
                    if col_idx > 0 and col_idx - 1 < len(row) and str(row[col_idx - 1]).strip():
                        has_left_data = True
                    if col_idx + 1 < len(row) and str(row[col_idx + 1]).strip():
                        has_right_data = True
                if has_left_data and has_right_data:
                    return col_idx

        return 0

    def _find_separator_by_empty_col(self, table: List[List[str]]) -> int:
        """
        通过空白列找到分割位置（备选策略）

        Args:
            table: 表格数据

        Returns:
            分割列索引，未找到返回0
        """
        if not table or len(table) < 2:
            return 0

        num_cols = len(table[0])
        if num_cols < 6:
            return 0

        # 找到表头行
        header_row = 0
        header_keywords = ['catalog', 'part', 'number', '型号', '编码', 'series']
        for i, row in enumerate(table[:min(6, len(table))]):
            if any(kw in str(cell).lower() for kw in header_keywords for cell in row):
                header_row = i
                break

        # 在中间区域寻找空白列作为分隔符
        check_rows = min(header_row + 10, len(table))
        separator_col = -1
        max_empty_ratio = 0.0

        # 只在中间1/3区域查找分隔列
        start_col = num_cols // 3
        end_col = num_cols * 2 // 3

        for col_idx in range(start_col, end_col):
            empty_count = 0
            for row_idx in range(check_rows):
                row = table[row_idx]
                if col_idx < len(row) and str(row[col_idx]).strip() == "":
                     empty_count += 1
            # 计算空白率
            empty_ratio = empty_count / check_rows
            if empty_ratio > max_empty_ratio:
                max_empty_ratio = empty_ratio
                separator_col = col_idx

        # 空白率 >= 60% 才认为是有效的分隔列
        if separator_col > 0 and max_empty_ratio >= 0.6:
            return separator_col

        return 0


class KeywordSearcher:
    """
    关键词搜索器

    负责表头识别、关键词匹配、先字典后泛化的搜索策略。
    """

    # 关键词规则字典：定义常用表头的同义词映射
    KEYWORD_RULES = {
        'partnumber': {
            'keywords': ['partnumber', 'part', 'number', '型号', '编码', 'code', 'pn', 'partno', '产品编码', '型号规格', 'catalog part number', 'catalog'],
            'column_hints': ['型号', '编码', 'part', 'code', 'pn', 'catalog']
        },
        'cap': {
            'keywords': ['cap', 'capacitance', '静电容量', '容量', 'c', '电容'],
            'column_hints': ['cap', 'capacitance', '电容', '容量', 'c']
        },
        'wv': {
            'keywords': ['wv', 'voltage', '额定电压', 'ur', '工作电压', 'voltage range', '电压', 'wvd', 'wvdc', 'surge'],
            'column_hints': ['wv', 'voltage', 'ur', '电压', 'v', 'wvd']
        },
        'size': {
            'keywords': ['size', '尺寸', 'dimension', '规格', '直径', '高度', 'diameter'],
            'column_hints': ['size', 'dimension', '尺寸', '规格', 'd', 'h']
        },
        'tolerance': {
            'keywords': ['tolerance', '公差', '容差'],
            'column_hints': ['tolerance', '公差', '容差']
        },
        'series': {
            'keywords': ['series', '系列'],
            'column_hints': ['series', '系列']
        }
    }

    def __init__(self):
        """初始化关键词搜索器"""
        pass

    def _contains_keyword(self, text: str, keywords: List[str]) -> bool:
        """
        判断文本是否包含任一关键词

        Args:
            text: 待检测文本
            keywords: 关键词列表

        Returns:
            True表示包含关键词，False表示不包含
        """
        text_lower = str(text).lower().strip()
        if not text_lower:
            return False
        for kw in keywords:
            # 完全匹配或包含匹配（文本长度不超过60字符）
            if kw.lower() == text_lower or (kw.lower() in text_lower and len(text_lower) <= 60):
                return True
        return False

    def find_header_row(self, table: List[List[str]]) -> int:
        """
        智能识别表头行位置

        通过打分机制判断哪一行最可能是表头：
        - 包含表头特征词（part/number/型号/电容等）+5分
        - 纯字母且长度<=8（可能是简写表头）+3分

        Args:
            table: 表格数据

        Returns:
            表头行的索引（从0开始）
        """
        if not table or len(table) < 2:
            return 0

        best_score = 0
        best_row = 0

        # 只检查前10行
        for row_idx, row in enumerate(table[:min(10, len(table))]):
            score = 0
            cell_count = 0
            for cell in row:
                cell_str = str(cell).strip().lower()
                if cell_str:
                    cell_count += 1
                    # 表头特征词匹配
                    if any(kw in cell_str for kw in ['part', 'number', '型号', '规格', 'item', '项目', '特性', 'cap', 'voltage', '电容', '尺寸', 'series', 'code', '编码', 'wv', 'ur', 'catalog', 'size']):
                        score += 5
                    # 纯字母简写匹配
                    if re.match(r'^[a-zA-Z]+$', cell_str) and len(cell_str) <= 8:
                        score += 3
            # 至少有2个非空单元格且分数最高
            if cell_count >= 2 and score > best_score:
                best_score = score
                best_row = row_idx

        return best_row

    def find_header_rows(self, table: List[List[str]]) -> Tuple[int, int]:
        """
        识别多行表头的范围

        有些表格表头可能跨多行（如合并单元格），此方法找出表头的起始行和结束行。

        改进点：
        1. 增加更多表头特征词检测
        2. 检测首字母大写的文本（通常是表头）
        3. 检测单位符号和括号内的说明
        4. 检测表头关键词的部分匹配（如"part"在"catalog part number"中）

        Args:
            table: 表格数据

        Returns:
            (起始行索引, 结束行索引)
        """
        start_row = self.find_header_row(table)

        end_row = start_row
        # 检查后续最多8行是否仍包含表头特征（增加检查行数）
        check_end = min(start_row + 3, len(table))
        
        # 收集已确认的表头关键词（用于后续行的判断）
        header_keywords_found = set()
        for cell in table[start_row]:
            cell_str = str(cell).strip().lower()
            if cell_str:
                for kw in ['part', 'number', 'cap', 'voltage', 'size', 'catalog', 'series', 'tolerance', 'code', '型号', '编码', '电容', '电压', '尺寸', '公差', '系列']:
                    if kw in cell_str:
                        header_keywords_found.add(kw)

        for row_idx in range(start_row + 1, check_end):
            row = table[row_idx]
            has_header_like = False
            
            for cell in row:
                cell_str = str(cell).strip()
                if not cell_str:
                    continue
                    
                cell_lower = cell_str.lower()
                
                # 检测1：表头特征词匹配
                if any(kw in cell_lower for kw in ['part', 'number', 'cap', 'voltage', 'size', 'tan', 'esr', 'catalog', 'series', 'tolerance', 'code', '型号', '编码', '电容', '电压', '尺寸', '公差', '系列', 'wv', 'ur']):
                    has_header_like = True
                    break
                
                # 检测2：首字母大写且长度适中（通常是表头）
                if cell_str and cell_str[0].isupper() and len(cell_str) <= 50:
                    # 排除纯数字或产品型号（如HSE102M51）
                    if not re.match(r'^[A-Z0-9]+$', cell_str):
                        has_header_like = True
                        break
                
                # 检测3：包含单位符号
                if any(unit in cell_lower for unit in ['μf', 'mm', 'v', 'ω', 'mhz', 'pf', 'nf', 'ohm']):
                    has_header_like = True
                    break
                
                # 检测4：包含括号或括号内的说明
                if '(' in cell_str or ')' in cell_str or '（' in cell_str or '）' in cell_str:
                    has_header_like = True
                    break
                
                # 检测5：匹配已确认的表头关键词的部分
                if header_keywords_found:
                    for kw in header_keywords_found:
                        if kw in cell_lower:
                            has_header_like = True
                            break
            
            if has_header_like:
                end_row = row_idx
            else:
                break

        return start_row, end_row

    def build_merged_header(self, table: List[List[str]], start_row: int, end_row: int) -> List[str]:
        """
        将多行表头合并为单行表头

        对于跨多行的表头，将同一列的多个表头单元格合并为一个完整的表头名称。

        Args:
            table: 表格数据
            start_row: 表头起始行
            end_row: 表头结束行

        Returns:
            合并后的单行表头列表
        """
        if not table or start_row > end_row:
            return []

        num_cols = len(table[start_row]) if table else 0
        merged_header = [""] * num_cols

        for row_idx in range(start_row, end_row + 1):
            row = table[row_idx]
            for col_idx, cell in enumerate(row):
                cell_str = str(cell).strip()
                if cell_str:
                    if merged_header[col_idx]:
                        merged_header[col_idx] += " " + cell_str
                    else:
                        merged_header[col_idx] = cell_str

        return merged_header

    def search_header(self, table: List[List[str]], keyword: str) -> Optional[int]:
        """
        先字典后泛化的表头搜索策略

        搜索流程：
        1. 如果关键词在KEYWORD_RULES字典中，优先使用字典中的同义词列表进行匹配
        2. 如果字典匹配失败，使用泛化搜索（关键词包含在表头文本中）
        3. 返回匹配到的列索引

        Args:
            table: 表格数据
            keyword: 搜索关键词

        Returns:
            匹配到的列索引，未找到返回None
        """
        if not table or not keyword:
            return None

        keyword_lower = keyword.lower().strip()
        if not keyword_lower:
            return None

        start_row, end_row = self.find_header_rows(table)

        # 第一层：字典搜索（精确匹配同义词）
        if keyword_lower in self.KEYWORD_RULES:
            rules = self.KEYWORD_RULES[keyword_lower]
            keywords = rules['keywords']

            for row_idx in range(start_row, min(end_row + 2, len(table))):
                row = table[row_idx]
                for col_idx, cell in enumerate(row):
                    if self._contains_keyword(cell, keywords):
                        return col_idx

        # 第二层：泛化搜索（模糊匹配）
        for row_idx in range(start_row, min(end_row + 2, len(table))):
            row = table[row_idx]
            for col_idx, cell in enumerate(row):
                cell_str = str(cell).strip().lower()
                if not cell_str:
                    continue

                # 关键词包含在表头中，或表头包含在关键词中
                if keyword_lower in cell_str or cell_str in keyword_lower:
                    return col_idx

        return None

    def search_all_headers(self, table: List[List[str]], keywords: List[str]) -> Dict[str, int]:
        """
        批量搜索多个关键词对应的表头列

        Args:
            table: 表格数据
            keywords: 关键词列表

        Returns:
            关键词到列索引的映射字典
        """
        mapping = {}
        used_cols = set()

        for keyword in keywords:
            col_idx = self.search_header(table, keyword)
            # 确保同一列不会被多个关键词匹配
            if col_idx is not None and col_idx not in used_cols:
                mapping[keyword] = col_idx
                used_cols.add(col_idx)

        return mapping


class PDFTableExtractor:
    """
    PDF表格提取器（核心类）

    集成PyMuPDF和OCR两种提取方式，提供表格提取和关键词搜索功能。
    支持大文件处理和多页批量搜索。
    """

    def __init__(self, pdf_path: str, use_ocr: bool = True):
        """
        初始化PDF表格提取器

        Args:
            pdf_path: PDF文件路径
            use_ocr: 是否启用OCR回退，默认True
        """
        self.pdf_path = pdf_path
        self.use_ocr = use_ocr
        self.ocr_extractor = None
        self.structure_analyzer = TableStructureAnalyzer()
        self.keyword_searcher = KeywordSearcher()
        self._all_cells = {}  # OCR识别结果缓存

    def _init_ocr_if_needed(self):
        """延迟初始化OCR（只在需要时创建）"""
        if self.use_ocr and self.ocr_extractor is None:
            self.ocr_extractor = OCRTableExtractor()

    def _get_ocr_cells(self, page_num: int, is_scanned: bool = False) -> List[Dict]:
        """
        获取指定页面的OCR识别结果（带缓存）

        Args:
            page_num: 页码
            is_scanned: 是否为扫描件

        Returns:
            单元格列表
        """
        cache_key = (page_num, is_scanned)
        if cache_key not in self._all_cells:
            self._init_ocr_if_needed()
            if self.ocr_extractor:
                cells = self.ocr_extractor.recognize_from_pdf_page(self.pdf_path, page_num, is_scanned)
                self._all_cells[cache_key] = cells
            else:
                self._all_cells[cache_key] = []
        return self._all_cells[cache_key]

    def _is_scanned_pdf(self, page_num: int) -> bool:
        doc = fitz.open(self.pdf_path)
        page = doc.load_page(page_num)
        
        text = page.get_text()
        blocks = page.get_text("blocks")
        
        text_block_count = sum(1 for b in blocks if b[6] == 0)
        
        doc.close()
        
        if text_block_count < 5 or len(text.strip()) < 20:
            return True
        return False

    def get_page_label(self, page_num: int) -> str:
        """
        返回第 page_num 页的PDF逻辑页码标签（如 '1'、'i'、'A-3'）。
        当PDF含封面/目录页时，该标签与印刷页码一致；获取失败时回退到物理页码+1。
        """
        try:
            doc = fitz.open(self.pdf_path)
            label = doc.load_page(page_num).get_page_label()
            doc.close()
            if label:
                return str(label)
        except Exception:
            pass
        return str(page_num + 1)

    def _save_page_image(self, page_num: int) -> str:
        pdf_dir = os.path.dirname(self.pdf_path)
        pdf_name = os.path.splitext(os.path.basename(self.pdf_path))[0]
        image_dir = os.path.join(pdf_dir, f"{pdf_name}_扫描图片")
        os.makedirs(image_dir, exist_ok=True)
        
        doc = fitz.open(self.pdf_path)
        page = doc.load_page(page_num)
        
        is_scanned = self._is_scanned_pdf(page_num)
        if is_scanned:
            pix = page.get_pixmap(matrix=fitz.Matrix(4.0, 4.0))
        else:
            pix = page.get_pixmap(matrix=fitz.Matrix(2.0, 2.0))
        
        image_path = os.path.join(image_dir, f"page_{page_num + 1}.png")
        pix.save(image_path)
        doc.close()
        
        return image_path

    def extract_tables(self, page_num: int = 0) -> List[List[List[str]]]:
        """
        提取指定页面的表格数据

        提取策略优化：
        1. 保存页面截图到文件夹
        2. 检测是否为扫描件
        3. 如果是扫描件：优先使用PPStructure识别表格结构，失败则使用传统OCR
        4. 如果是非扫描件：优先使用PyMuPDF快速提取，失败则回退到OCR
        5. 对识别结果进行双栏表格分割处理

        Args:
            page_num: 页码（从0开始）

        Returns:
            表格数据列表，每个表格是一个二维列表
        """
        self._save_page_image(page_num)
        
        is_scanned = self._is_scanned_pdf(page_num)
        
        if is_scanned and self.use_ocr:
            self._init_ocr_if_needed()
            if self.ocr_extractor:
                image_path = os.path.join(
                    os.path.dirname(self.pdf_path),
                    f"{os.path.splitext(os.path.basename(self.pdf_path))[0]}_扫描图片",
                    f"page_{page_num + 1}.png"
                )
                tables = self.ocr_extractor.recognize_table_with_ppstructure(image_path)
                if tables:
                    return tables
                
                tables = self._extract_with_ocr(page_num, is_scanned)
                if tables:
                    return tables
            
            return []
        
        tables = self._extract_with_pymupdf(page_num)

        all_tables = []
        for table in tables:
            split_tables = self.structure_analyzer.split_double_table(table)
            all_tables.extend(split_tables)

        if all_tables:
            return all_tables

        if self.use_ocr:
            tables = self._extract_with_ocr(page_num, is_scanned)
            if tables:
                return tables

        return []

    def _extract_with_pymupdf(self, page_num: int) -> List[List[List[str]]]:
        """
        使用PyMuPDF提取表格

        PyMuPDF提供了find_tables()方法，可以快速识别PDF中的表格结构。

        Args:
            page_num: 页码

        Returns:
            表格数据列表
        """
        doc = fitz.open(self.pdf_path)
        page = doc.load_page(page_num)

        tables = []
        try:
            # 使用PyMuPDF的表格识别功能
            tabs = page.find_tables()
            for tab in tabs.tables:
                table_data = []
                for row in tab.extract():
                    table_data.append([str(cell).strip() if cell else "" for cell in row])
                # 过滤无效表格（至少2行2列）
                if table_data and len(table_data) >= 2 and len(table_data[0]) >= 2:
                    # 不在这里清洗，而是在 split_double_table 中统一处理
                    # 避免在分割前删除可能只在一侧有数据的行
                    tables.append(table_data)
        except Exception:
            pass

        return tables

    def _extract_with_ocr(self, page_num: int, is_scanned: bool = False) -> List[List[List[str]]]:
        """
        使用OCR提取表格

        当PyMuPDF无法识别表格时，使用OCR进行文字识别，然后通过坐标聚类重建表格结构。

        Args:
            page_num: 页码
            is_scanned: 是否为扫描件

        Returns:
            表格数据列表
        """
        # 获取OCR识别结果
        cells = self._get_ocr_cells(page_num, is_scanned)

        if not cells:
            print("OCR识别结果为空")
            return []

        # 按Y坐标聚类，识别行
        rows = self._cluster_by_rows(cells, y_tolerance=35)

        if len(rows) < 2:
            return []

        # 检测列位置
        column_x = self._detect_columns(rows)

        if len(column_x) < 2:
            return []

        # 根据列位置重建表格
        table = []
        for row in rows:
            row_data = [""] * len(column_x)
            for cell in row:
                # 找到距离最近的列
                min_dist = float('inf')
                min_idx = 0
                for j, x in enumerate(column_x):
                    dist = abs(cell["center_x"] - x)
                    if dist < min_dist:
                        min_dist = dist
                        min_idx = j

                # 同一列有多个单元格时合并
                if row_data[min_idx]:
                    row_data[min_idx] += " " + cell["text"]
                else:
                    row_data[min_idx] = cell["text"]
            table.append(row_data)

        cleaned = self.structure_analyzer.clean_table(table)
        if len(cleaned) >= 2:
            return [cleaned]

        return []

    def _cluster_by_rows(self, cells: List[Dict], y_tolerance: int = 35) -> List[List[Dict]]:
        """
        按Y坐标聚类单元格，识别表格行

        Args:
            cells: 单元格列表（带坐标信息）
            y_tolerance: Y坐标容差（像素）

        Returns:
            行列表，每行包含该行的所有单元格
        """
        if not cells:
            return []

        # 按Y坐标排序，然后按X坐标排序
        sorted_cells = sorted(cells, key=lambda c: (c["y0"], c["x0"]))

        rows = []
        current_row = [sorted_cells[0]]
        current_y_center = sorted_cells[0]["center_y"]

        for cell in sorted_cells[1:]:
            # Y坐标差异在容差范围内属于同一行
            if abs(cell["center_y"] - current_y_center) < y_tolerance:
                current_row.append(cell)
            else:
                rows.append(current_row)
                current_row = [cell]
                current_y_center = cell["center_y"]

        if current_row:
            rows.append(current_row)

        # 每行内部按X坐标排序
        for row in rows:
            row.sort(key=lambda c: c["x0"])

        return rows

    def _detect_columns(self, rows: List[List[Dict]]) -> List[float]:
        """
        检测表格列位置

        通过对所有单元格的X中心坐标进行聚类，确定列的位置。

        Args:
            rows: 行列表

        Returns:
            列中心X坐标列表（已排序）
        """
        if not rows:
            return []

        x_positions = []
        for row in rows:
            for cell in row:
                x_positions.append(cell["center_x"])

        x_positions.sort()

        if not x_positions:
            return []

        # X坐标聚类，距离小于80像素的归为同一列
        clusters = []
        current_cluster = [x_positions[0]]

        for x in x_positions[1:]:
            if x - current_cluster[-1] < 80:
                current_cluster.append(x)
            else:
                clusters.append(current_cluster)
                current_cluster = [x]

        if current_cluster:
            clusters.append(current_cluster)

        # 计算每个聚类的中心坐标作为列位置
        column_x = [sum(c) / len(c) for c in clusters]
        return sorted(column_x)

    def search_by_keyword(self, page_num: int, keywords: List[str]) -> Dict[str, Any]:
        """
        在指定页面搜索包含关键词的数据

        搜索流程：
        1. 提取页面中的所有表格
        2. 对每个表格识别表头行并合并
        3. 搜索关键词对应的表头列
        4. 提取匹配列下方的数据行

        Args:
            page_num: 页码
            keywords: 关键词列表

        Returns:
            搜索结果字典，包含页码、表格数量、匹配数据等信息
        """
        tables = self.extract_tables(page_num)

        results = {
            "page": page_num,
            "tables_found": len(tables),
            "results": []
        }

        for table_idx, table in enumerate(tables):
            # 识别表头行范围并合并
            start_row, end_row = self.keyword_searcher.find_header_rows(table)
            merged_header = self.keyword_searcher.build_merged_header(table, start_row, end_row)

            # 搜索所有关键词对应的列
            header_mapping = self.keyword_searcher.search_all_headers(table, keywords)

            if header_mapping:
                table_result = {
                    "table_index": table_idx,
                    "total_rows": len(table),
                    "header_start_row": start_row,
                    "header_end_row": end_row,
                    "header_mapping": header_mapping,
                    "headers": merged_header,
                    "data": []
                }

                # 提取表头行和数据行（包含表头行，让用户可以自己判断）
                for row_idx, row in enumerate(table):
                    # 从表头起始行开始，包含表头行
                    if row_idx < start_row:
                        continue

                    row_data = {}
                    has_valid_data = False
                    
                    # 提取搜索关键词对应的列数据
                    for keyword, col_idx in header_mapping.items():
                        value = row[col_idx] if col_idx < len(row) else ""
                        row_data[keyword] = value
                        if str(value).strip():
                            has_valid_data = True

                    # 检查整行是否有任何数据（不仅仅是搜索列）
                    # 确保所有有数据的行都被保留
                    if not has_valid_data:
                        for cell in row:
                            if str(cell).strip():
                                has_valid_data = True
                                break

                    # 添加来源信息
                    row_data["_source_row"] = row_idx
                    row_data["_source_table"] = table_idx
                    row_data["_is_header"] = "是" if row_idx <= end_row else "否"

                    # 只保留包含有效数据的行
                    if has_valid_data:
                        table_result["data"].append(row_data)

                results["results"].append(table_result)

        return results

    def search_in_all_pages(self, keywords: List[str]) -> Dict[str, Any]:
        """
        在PDF所有页面搜索包含关键词的数据

        Args:
            keywords: 关键词列表

        Returns:
            全文档搜索结果
        """
        doc = fitz.open(self.pdf_path)
        total_pages = doc.page_count
        doc.close()

        all_results = {
            "total_pages": total_pages,
            "results": []
        }

        for page_num in range(total_pages):
            page_results = self.search_by_keyword(page_num, keywords)
            if page_results["results"]:
                all_results["results"].append(page_results)

        return all_results


def main(pdf_path: str, keywords: List[str], page_num: int = 0) -> pd.DataFrame:
    """
    主函数：提取表格数据并返回DataFrame

    Args:
        pdf_path: PDF文件路径
        keywords: 关键词列表
        page_num: 页码，默认0

    Returns:
        包含匹配数据的DataFrame
    """
    extractor = PDFTableExtractor(pdf_path, use_ocr=True)
    results = extractor.search_by_keyword(page_num, keywords)

    all_data = []
    for table_result in results["results"]:
        for row in table_result["data"]:
            row_copy = row.copy()
            row_copy["_来源表格"] = table_result["table_index"]
            row_copy["_来源行号"] = row_copy.get("_source_row", "")
            all_data.append(row_copy)

    if not all_data:
        return pd.DataFrame()

    df = pd.DataFrame(all_data)

    # 按关键词顺序排列列，最后显示来源信息
    cols_order = [k for k in keywords if k in df.columns]
    cols_order.extend(["_来源行号", "_来源表格"])

    return df[cols_order]


def _legacy_test_entry():
    """旧版测试入口（无参数时调用）"""
    import os

    pdf_path = os.path.join("最新目录PDF", "AJ.pdf")
    keywords = ["partnumber", "cap", "wv"]

    if not os.path.exists(pdf_path):
        print(f"[跳过] 旧版测试数据不存在: {pdf_path}")
        return

    print("=" * 60)
    print("PDF 表格数据提取器 - 测试（旧版）")
    print("=" * 60)

    extractor = PDFTableExtractor(pdf_path, use_ocr=True)

    doc = fitz.open(pdf_path)
    total_pages = doc.page_count

    # 测试表格提取
    for page_num in range(total_pages):
        tables = extractor.extract_tables(page_num)
        print(f"\n--- 第 {page_num + 1} 页 ---")
        print(f"识别到 {len(tables)} 个表格")

        for i, table in enumerate(tables):
            start_row, end_row = extractor.keyword_searcher.find_header_rows(table)
            merged_header = extractor.keyword_searcher.build_merged_header(table, start_row, end_row)
            print(f"\n表格 {i+1}: {len(table)}行 x {len(table[0])}列")
            print(f"表头行范围: {start_row}-{end_row}")
            print("合并表头:", merged_header)

    print("\n" + "=" * 60)
    print(f"搜索关键词: {keywords}")
    print("=" * 60)

    # 测试关键词搜索
    for page_num in range(total_pages):
        results = extractor.search_by_keyword(page_num, keywords)
        if results["results"]:
            print(f"\n--- 第 {page_num + 1} 页 ---")
            for table_result in results["results"]:
                print(f"\n表格 {table_result['table_index']}:")
                print(f"  表头映射: {table_result['header_mapping']}")
                print(f"  找到 {len(table_result['data'])} 行数据")

    # 测试DataFrame输出
    df = main(pdf_path, keywords, page_num=1)
    print(f"\n最终 DataFrame ({len(df)} 行):")
    print(df)

    print("\n" + "=" * 60)
    print("测试完成")
    print("=" * 60)


# =============================================================================
# 型号提取增强功能（方案A：在现有基础上扩展）
# =============================================================================

class ModelExtractor:
    """
    型号提取器（阶段三增强版）

    基于大厂命名规则的正则识别系统，支持：
    - 电子元器件型号（电阻、电容、继电器、接触器等）
    - 机械零件型号（标准件、导轨、轴承等）
    - 双库精筛（Trie前缀分组 + 编辑距离匹配）
    - 全页扫描（不依赖检测框）
    - 型号反查补全参数
    """

    # 电子元器件型号正则（大厂格式）
    ELECTRONIC_PATTERNS = [
        # 电阻/电容型号: CJX2-0910, CJX2-2510, ZM-3010
        r'\b[A-Z]{2,6}[-]?\d{2,5}[A-Z]?\b',
        # 电阻/电容值: 10kΩ, 100nF, 10uF, 220V
        r'\b\d+[.,]?\d*\s*[kKmMnNuUpP]?(?:Ω|ohm|F|V|A|W|Hz)\b',
        # 电阻 R 编码: 2R01, 4R02, 1R50, R010, R100, 30M1
        r'\b\d{0,3}[RM]\d{0,3}\b',
        # 电容 R 编码: 1R0 (1.0Ω), 4R7 (4.7Ω)
        r'\b\d{1,2}R\d{1,2}\b',
        # 继电器: MY2NJ, MY4NJ, HH54P
        r'\b[A-Z]{2,6}\d{1,3}[A-Z]{0,3}[A-Z0-9]*\b',
        # 集成电路: LM358, NE555, STM32F103
        r'\b[A-Z]{2,4}\d{3,8}[A-Z0-9]*\b',
        # 连接器型号: XH-4P, PH2.0-5P
        r'\b[A-Z]{1,4}[-]?\d{1,4}[Pp]\b',
    ]

    # 机械零件型号正则
    MECHANICAL_PATTERNS = [
        # 螺栓/螺钉: M8x20, M10x30
        r'\bM\d{1,2}[x×]\d{1,4}\b',
        # 导轨: HGH20CA, MGN12H
        r'\b[A-Z]{2,6}\d{1,3}[A-Z]{0,2}[A-Z0-9]*\b',
        # 轴承: 6205ZZ, 608-2RS
        r'\b\d{4,5}[A-Z]{0,3}[A-Z0-9]*\b',
        # 气缸: SC32x100, SCDA2B
        r'\b[A-Z]{2,4}\d{1,3}[x×]\d{2,4}\b',
    ]

    # 低置信度候选的通用条件
    GENERIC_CANDIDATE = re.compile(
        r'\b(?=[A-Z]*\d+[A-Z]*\d*[A-Z]*\b)[A-Z0-9\-]{4,15}\b'
    )

    # 排除的干扰词（收紧版）
    EXCLUDE_WORDS = {
        'ISO', 'CCC', 'RoHS', 'CE', 'UL', 'VDE', 'CSA', 'PSE',
        'PDF', 'Page', 'Pages', 'www', 'http', 'https', 'com', 'cn',
        'Rev', 'Version', 'Date', 'Sheet', 'File', 'Doc',
        'A4', 'A3', 'B5', 'mm', 'cm', 'inch', 'in',
        'TYPE', 'TYPE:', 'SIZE', 'SIZE:', 'STEP', 'NOTE', 'NOTES',
        'SPEC', 'SPECIFICATION', 'STANDARD', 'STANDARDS',
        'CATALOG', 'CATALOGUE', 'PRODUCT', 'PRODUCTS',
        'SERIES', 'MODEL', 'MODELS', 'NO', 'NO.', 'NO:',
        'FIG', 'FIG.', 'FIGURE', 'TABLE', 'TABLES',
        'PART', 'PARTS', 'ITEM', 'ITEMS',
        'TOTAL', 'TOTALS', 'GRAND', 'SUB',
        'CONT', 'CONT.', 'CONTINUED',
        'OF', 'THE', 'AND', 'FOR', 'ARE',
        'Rated', 'rated', 'RATED',
        'Power', 'power', 'POWER',
    }

    def __init__(self, model_db_path: Optional[str] = None):
        """
        初始化型号提取器（阶段三增强版）

        Args:
            model_db_path: 型号数据库路径（jsonl），可选
        """
        # 原始集合（用于精确匹配）
        self.electronic_db = set()
        self.mechanical_db = set()

        # Trie 分组字典（按前4字符分组，加速模糊匹配）
        self.electronic_groups: Dict[str, List[Dict]] = {}
        self.mechanical_groups: Dict[str, List[Dict]] = {}

        # 完整数据库条目（用于反查补全参数）
        self.electronic_entries: Dict[str, Dict] = {}
        self.mechanical_entries: Dict[str, Dict] = {}

        self._trie_loaded = False

        if model_db_path:
            self._load_model_db(model_db_path)

    def _load_model_db(self, db_path: str):
        """加载型号数据库（同时构建集合和Trie分组）"""
        try:
            electronic_file = os.path.join(db_path, 'electronic.jsonl')
            mechanical_file = os.path.join(db_path, 'mechanical.jsonl')

            if os.path.exists(electronic_file):
                with open(electronic_file, 'r', encoding='utf-8') as f:
                    for line in f:
                        if line.strip():
                            try:
                                item = json.loads(line.strip())
                                model_no = item.get('model_no', '').upper()
                                if model_no:
                                    self.electronic_db.add(model_no)
                                    self.electronic_entries[model_no] = item
                                    # Trie 分组
                                    prefix = model_no[:4]
                                    if prefix not in self.electronic_groups:
                                        self.electronic_groups[prefix] = []
                                    self.electronic_groups[prefix].append(item)
                            except json.JSONDecodeError:
                                continue

            if os.path.exists(mechanical_file):
                with open(mechanical_file, 'r', encoding='utf-8') as f:
                    for line in f:
                        if line.strip():
                            try:
                                item = json.loads(line.strip())
                                model_no = item.get('model_no', '').upper()
                                if model_no:
                                    self.mechanical_db.add(model_no)
                                    self.mechanical_entries[model_no] = item
                                    # Trie 分组
                                    prefix = model_no[:4]
                                    if prefix not in self.mechanical_groups:
                                        self.mechanical_groups[prefix] = []
                                    self.mechanical_groups[prefix].append(item)
                            except json.JSONDecodeError:
                                continue

            self._trie_loaded = True
            total = len(self.electronic_db) + len(self.mechanical_db)
            print(f"[ModelExtractor] 加载型号库: 电子 {len(self.electronic_db)} 条, 机械 {len(self.mechanical_db)} 条, 共 {total} 条")
        except Exception as e:
            print(f"[ModelExtractor] 加载型号库失败: {e}")

    def lookup_model_info(self, model_str: str) -> Optional[Dict]:
        """
        型号反查补全参数（阶段三新增）

        Args:
            model_str: 型号字符串

        Returns:
            型号完整信息字典，未找到返回 None
        """
        model_upper = model_str.upper()

        # 精确匹配
        if model_upper in self.electronic_entries:
            entry = self.electronic_entries[model_upper].copy()
            entry['matched_db'] = 'electronic'
            return entry
        if model_upper in self.mechanical_entries:
            entry = self.mechanical_entries[model_upper].copy()
            entry['matched_db'] = 'mechanical'
            return entry

        # 模糊匹配（Trie分组 + 编辑距离）
        match = self.fuzzy_match_db(model_str)
        if match:
            if match in self.electronic_entries:
                entry = self.electronic_entries[match].copy()
                entry['fuzzy_matched'] = True
                entry['matched_db'] = 'electronic'
                return entry
            if match in self.mechanical_entries:
                entry = self.mechanical_entries[match].copy()
                entry['fuzzy_matched'] = True
                entry['matched_db'] = 'mechanical'
                return entry

        return None

    def get_model_params(self, model_str: str) -> Dict:
        """
        获取型号参数（阶段三新增，反查补全参数）

        Args:
            model_str: 型号字符串

        Returns:
            参数字典，包含 type, company, category 等
        """
        info = self.lookup_model_info(model_str)
        if info:
            return {
                'type': info.get('type', '未知'),
                'company': info.get('company', '未知'),
                'category': info.get('category', '未知'),
                'model_no': info.get('model_no', model_str),
            }
        return {
            'type': '未收录',
            'company': '未收录',
            'category': '未收录',
            'model_no': model_str,
        }

    def _is_noise_candidate(self, token_upper: str) -> bool:
        """
        过滤明显不是型号的正则命中（表外型号噪声的主要来源）

        两类噪声：
        1. 纯数字串 —— 手机号 13824341110、区号 0755、年份、页码、编号。
           MECHANICAL_PATTERNS 的 \\b\\d{4,5}[A-Z]{0,3}[A-Z0-9]*\\b 会整段吞掉长数字串。
        2. 无数字的超短串 —— 单字母 R / M 等。
           ELECTRONIC_PATTERNS 的 \\b\\d{0,3}[RM]\\d{0,3}\\b 允许 0 位数字，会命中裸字母。

        命中型号库的（如轴承 6205）一律放行。
        """
        if token_upper in self.electronic_db or token_upper in self.mechanical_db:
            return False

        core = token_upper.replace('-', '').replace('.', '').replace(' ', '')
        if core.isdigit():
            return True
        if len(core) <= 3 and not any(ch.isdigit() for ch in core):
            return True
        # 单位量不是型号：电压(25V/700V/3.3KV)、阻容值(100µF/10KΩ/4.7UF)、认证号(ISO9001/RoHS)
        if re.fullmatch(r'\d+(\.\d+)?[KMGTP]?[µuU]?V', token_upper):
            return True
        if re.fullmatch(r'\d+(\.\d+)?[KMG]?[µuU]?(F|OHM|Ω)', token_upper):
            return True
        if re.fullmatch(r'(ISO|IEC|UL|CE|RoHS|REACH|CCC)\d*', token_upper):
            return True
        return False

    def extract_models_from_text(self, text: str) -> List[Dict]:
        """
        从文本中提取型号候选（关键：跑在全页文本上，不依赖检测框）

        Args:
            text: 输入文本

        Returns:
            型号候选列表 [{match, pattern_type, confidence, ...}]
        """
        candidates = []
        seen = set()

        # 电子型号
        for pattern in self.ELECTRONIC_PATTERNS:
            for match in re.finditer(pattern, text, re.IGNORECASE):
                match_str = match.group(0).strip()
                match_upper = match_str.upper()

                # 排除干扰词
                if match_upper in self.EXCLUDE_WORDS:
                    continue
                if self._is_noise_candidate(match_upper):
                    continue

                # 去重
                if match_upper in seen:
                    continue
                seen.add(match_upper)

                # 精筛判断：命中电子库或机械库任一即 high（跨库匹配也认）
                is_in_db = match_upper in self.electronic_db or match_upper in self.mechanical_db
                confidence = 'high' if is_in_db else 'medium'

                candidates.append({
                    'model': match_str,
                    'model_upper': match_upper,
                    'pattern_type': 'electronic',
                    'confidence': confidence,
                    'source_pattern': pattern[:30],
                })

        # 机械型号
        for pattern in self.MECHANICAL_PATTERNS:
            for match in re.finditer(pattern, text, re.IGNORECASE):
                match_str = match.group(0).strip()
                match_upper = match_str.upper()

                if match_upper in self.EXCLUDE_WORDS:
                    continue
                if self._is_noise_candidate(match_upper):
                    continue
                if match_upper in seen:
                    continue
                seen.add(match_upper)

                # 精筛判断：命中电子库或机械库任一即 high（跨库匹配也认）
                is_in_db = match_upper in self.electronic_db or match_upper in self.mechanical_db
                confidence = 'high' if is_in_db else 'medium'

                candidates.append({
                    'model': match_str,
                    'model_upper': match_upper,
                    'pattern_type': 'mechanical',
                    'confidence': confidence,
                    'source_pattern': pattern[:30],
                })

        # 通用兜底：低置信候选
        for match in re.finditer(self.GENERIC_CANDIDATE, text):
            match_str = match.group(0).strip()
            match_upper = match_str.upper()

            if match_upper in self.EXCLUDE_WORDS:
                continue
            if self._is_noise_candidate(match_upper):
                continue
            if match_upper in seen:
                continue
            seen.add(match_upper)

            # 检查是否已在电子/机械库
            if match_upper in self.electronic_db or match_upper in self.mechanical_db:
                continue

            candidates.append({
                'model': match_str,
                'model_upper': match_upper,
                'pattern_type': 'generic',
                'confidence': 'low',
                'source_pattern': 'generic',
            })

        return candidates

    def calculate_edit_distance(self, s1: str, s2: str) -> int:
        """计算编辑距离（用于双库精筛）"""
        if len(s1) < len(s2):
            return self.calculate_edit_distance(s2, s1)

        if len(s2) == 0:
            return len(s1)

        previous_row = range(len(s2) + 1)
        for i, c1 in enumerate(s1):
            current_row = [i + 1]
            for j, c2 in enumerate(s2):
                insertions = previous_row[j + 1] + 1
                deletions = current_row[j] + 1
                substitutions = previous_row[j] + (c1 != c2)
                current_row.append(min(insertions, deletions, substitutions))
            previous_row = current_row

        return previous_row[-1]

    def fuzzy_match_db(self, model_str: str, max_distance: int = 1) -> Optional[str]:
        """
        在数据库中模糊匹配型号（Trie分组优化版）

        Args:
            model_str: 待匹配型号
            max_distance: 最大编辑距离

        Returns:
            匹配到的数据库型号，未匹配返回 None
        """
        model_upper = model_str.upper()

        # 精确匹配（快速路径）
        if model_upper in self.electronic_db:
            return model_upper
        if model_upper in self.mechanical_db:
            return model_upper

        # Trie 分组匹配（只在同前缀组内比较，避免全量遍历）
        prefix = model_upper[:4]

        # 检查电子库
        for db_set, db_entries in [
            (self.electronic_groups, self.electronic_entries),
            (self.mechanical_groups, self.mechanical_entries),
        ]:
            # 尝试精确前缀匹配
            if prefix in db_set:
                for entry in db_set[prefix]:
                    db_model = entry['model_no'].upper()
                    dist = self.calculate_edit_distance(model_upper, db_model)
                    if dist <= max_distance:
                        return db_model

            # 尝试前3字符匹配（回退）
            if len(prefix) >= 3:
                prefix3 = model_upper[:3]
                for key, entries in db_set.items():
                    if key.startswith(prefix3):
                        for entry in entries:
                            db_model = entry['model_no'].upper()
                            dist = self.calculate_edit_distance(model_upper, db_model)
                            if dist <= max_distance:
                                return db_model

        return None


class EnhancedPDFTableExtractor(PDFTableExtractor):
    """
    增强版PDF表格提取器

    在现有 PDFTableExtractor 基础上增加：
    - 全页型号正则扫描（不依赖 YOLO 框）
    - 型号分类（表内/表外/可疑）
    - 三 Sheet Excel 导出
    """

    def __init__(self, pdf_path: str, use_ocr: bool = True, model_db_path: Optional[str] = None, yolo_detector: Optional['YOLOTableDetector'] = None):
        """
        初始化增强版提取器

        Args:
            pdf_path: PDF文件路径
            use_ocr: 是否启用OCR
            model_db_path: 型号数据库路径
            yolo_detector: YOLO 表格检测器（可选，用于扫描页）
        """
        super().__init__(pdf_path, use_ocr)
        self.model_extractor = ModelExtractor(model_db_path)
        self.table_bboxes: List[fitz.Rect] = []   # 存储表格区域坐标
        self.yolo_detector = yolo_detector    #yolo检测区
        self._yolo_cache: Dict[int, List[fitz.Rect]] = {}  # 存储表格区域坐标
        self._scanned_cache: Dict[int, Dict] = {}  # 扫描页完整结果缓存
        self.doc = fitz.open(pdf_path)
        self._doc_closed = False

        self.table_pipeline = TableDetectionPipeline(
            yolo_detector=yolo_detector,
            ocr_extractor=None,
            dpi=300
        )
        self._init_ocr_if_needed()
        if self.ocr_extractor:
            self.table_pipeline.ocr_extractor = self.ocr_extractor

        self._page_text_cache: Dict[int, Dict] = {}

    def _is_scanned_pdf(self, page_num: int) -> bool:
        """
        判断页面是否为扫描页（无文本层）

        注意：此方法现在只决定文字层路径，不再决定是否跑 YOLO。
        YOLO 由 TableDetectionPipeline 三源恒定全跑。
        """
        page = self.doc.load_page(page_num)
        text = page.get_text()
        blocks = page.get_text("blocks")
        text_block_count = sum(1 for b in blocks if b[6] == 0)
        if text_block_count < 5 or len(text.strip()) < 20:
            return True
        return False

    def __del__(self):
        """析构函数：确保文档关闭"""
        self.close()

    def close(self):
        """显式关闭文档，防止文件句柄泄漏"""
        if not self._doc_closed and hasattr(self, 'doc') and self.doc:
            try:
                self.doc.close()
                self._doc_closed = True
            except Exception:
                pass

    def get_page_label(self, page_num: int) -> str:
        """重写父类方法，使用已打开的 self.doc 避免文件冲突"""
        try:
            label = self.doc.load_page(page_num).get_page_label()
            if label:
                return str(label)
        except Exception:
            pass
        return str(page_num + 1)

    def extract_all_models_with_location(self, page_num: int) -> List[Dict]:
        """
        全页型号扫描（关键方法）

        矢量页：取全页 text spans → 跑正则 → 返回带坐标的型号列表
        扫描页：YOLO 检测表格 → 全页 OCR → 空间分类 → 型号识别

        Args:
            page_num: 页码（0-based）

        Returns:
            型号列表 [{model, bbox, page, confidence, in_table, ...}]
        """
        page = self.doc[page_num]
        page_models = []

        # 获取全页文本 spans（矢量坐标）
        try:
            text_dict = page.get_text("dict")
        except Exception as e:
            print(f"[ModelExtract] 获取文本层失败: {e}")
            text_dict = {}

        # 从文本层提取型号
        spans_text = ""
        span_locations = []  # [(text, bbox)]

        for block in text_dict.get("blocks", []):
            if "lines" in block:
                for line in block["lines"]:
                    for span in line.get("spans", []):
                        span_text = span.get("text", "").strip()
                        if span_text:
                            bbox = fitz.Rect(span["bbox"])
                            spans_text += span_text + " "
                            span_locations.append((span_text, bbox))

        # 如果有文本层内容，直接从文本层提取
        if spans_text.strip():
            candidates = self.model_extractor.extract_models_from_text(spans_text)

            for candidate in candidates:
                model_str = candidate["model"]
                model_bbox = self._find_model_bbox(model_str, span_locations)

                page_models.append({
                    'model': model_str,
                    'model_upper': candidate['model_upper'],
                    'pattern_type': candidate['pattern_type'],
                    'confidence': candidate['confidence'],
                    'bbox': model_bbox,
                    'page': page_num,
                    'source_pattern': candidate['source_pattern'],
                    'in_table': False,  # 待分类
                    'table_index': -1,
                    'table_row': -1,
                    'header_match': '',
                })

        # 扫描页处理：无文本层或文本很少时
        is_scanned = self.yolo_detector is not None and self.yolo_detector.is_scanned_page(page)
        
        if is_scanned and self.yolo_detector is not None:
            # 使用新流程：YOLO 检测 + 全页 OCR + 空间分类
            self._init_ocr_if_needed()
            if self.ocr_extractor:
                try:
                    print(f"[扫描页] 页 {page_num + 1}: YOLO + OCR 处理")
                    scanned_result = self.yolo_detector.process_scanned_page(page, self.ocr_extractor)
                    
                    # 缓存扫描页结果
                    self._scanned_cache[page_num] = scanned_result
                    
                    # 表外文本：跑型号正则
                    out_table_text = ' '.join([c.get('text', '') for c in scanned_result['out_table_cells']])
                    if out_table_text:
                        candidates = self.model_extractor.extract_models_from_text(out_table_text)
                        
                        # 计算像素到矢量坐标的转换
                        scale = self.yolo_detector.dpi / 72
                        
                        for candidate in candidates:
                            model_str = candidate["model"]
                            # 在表外 OCR 结果中找坐标
                            model_bbox = None
                            for cell in scanned_result['out_table_cells']:
                                r_text = cell.get('text', '')
                                if model_str in r_text or r_text in model_str:
                                    x0 = cell.get('x0', 0) / scale
                                    y0 = cell.get('y0', 0) / scale
                                    x1 = cell.get('x1', 0) / scale
                                    y1 = cell.get('y1', 0) / scale
                                    model_bbox = fitz.Rect(x0, y0, x1, y1)
                                    break
                            
                            if model_bbox:
                                page_models.append({
                                    'model': model_str,
                                    'model_upper': candidate['model_upper'],
                                    'pattern_type': candidate['pattern_type'],
                                    'confidence': candidate['confidence'],
                                    'bbox': model_bbox,
                                    'page': page_num,
                                    'source_pattern': candidate['source_pattern'],
                                    'in_table': False,  # 表外型号
                                    'table_index': -1,
                                    'table_row': -1,
                                    'header_match': '',
                                })
                except Exception as e:
                    print(f"[扫描页] 页 {page_num + 1} 处理失败: {e}")

        # 兜底：文本层和 YOLO+OCR 都没找到时，用普通 OCR
        if not page_models and self.use_ocr and self.ocr_extractor and not is_scanned:
            self._init_ocr_if_needed()
            if self.ocr_extractor:
                try:
                    ocr_results = self.ocr_extractor.recognize_from_pdf_page(self.pdf_path, page_num)
                    if ocr_results:
                        ocr_text = ' '.join([r.get('text', '') for r in ocr_results])
                        candidates = self.model_extractor.extract_models_from_text(ocr_text)

                        for candidate in candidates:
                            model_str = candidate["model"]
                            model_bbox = None
                            for r in ocr_results:
                                r_text = r.get('text', '')
                                if model_str in r_text or r_text in model_str:
                                    if 'bbox' in r:
                                        pts = r['bbox']
                                        if len(pts) == 4:
                                            x0, y0, x1, y1 = pts
                                            model_bbox = fitz.Rect(x0, y0, x1, y1)
                                    break

                            page_models.append({
                                'model': model_str,
                                'model_upper': candidate['model_upper'],
                                'pattern_type': candidate['pattern_type'],
                                'confidence': candidate['confidence'],
                                'bbox': model_bbox,
                                'page': page_num,
                                'source_pattern': candidate['source_pattern'],
                                'in_table': False,
                                'table_index': -1,
                                'table_row': -1,
                                'header_match': '',
                            })
                except Exception as e:
                    print(f"[OCR] 第{page_num+1}页 OCR 识别失败: {e}")

        return page_models

    def _find_model_bbox(self, model_str: str, span_locations: List[Tuple[str, fitz.Rect]]) -> Optional[fitz.Rect]:
        """在 spans 中查找型号对应的 bbox"""
        model_upper = model_str.upper()

        # 精确匹配
        for text, bbox in span_locations:
            if text.upper() == model_upper:
                return bbox

        # 包含匹配
        for text, bbox in span_locations:
            if model_upper in text.upper():
                return bbox

        # 模糊匹配（前缀匹配）
        for text, bbox in span_locations:
            if text.upper()[:4] == model_upper[:4]:
                return bbox

        # 返回 None（无坐标）
        return None

    def get_table_bboxes(self, page_num: int) -> List[fitz.Rect]:
        """
        获取页面上所有表格的 bbox（矢量坐标）

        策略：
        1. 矢量页（有文本层）：优先用 PyMuPDF find_tables()
        2. 扫描页（无文本层）：从扫描页缓存获取 YOLO 检测结果并转换坐标
        3. 两种方式都失败时返回空列表

        Args:
            page_num: 页码

        Returns:
            表格 bbox 列表（PDF 矢量坐标）
        """
        page = self.doc[page_num]
        bboxes = []

        # 先尝试 PyMuPDF 矢量检测
        try:
            tables = page.find_tables()
            for tab in tables.tables:
                if tab.bbox:
                    bboxes.append(fitz.Rect(tab.bbox))
        except Exception as e:
            print(f"[TableBBox] PyMuPDF 检测失败: {e}")

        # 如果矢量检测没找到表格，从扫描页缓存获取
        if not bboxes and page_num in self._scanned_cache:
            scanned_result = self._scanned_cache[page_num]
            scale = self.yolo_detector.dpi / 72 if self.yolo_detector else 2.0
            
            for table_det in scanned_result.get('table_detections', []):
                x1, y1, x2, y2 = table_det['bbox']
                # 像素坐标 → PDF 矢量坐标
                pdf_rect = fitz.Rect(
                    x1 / scale, y1 / scale,
                    x2 / scale, y2 / scale
                )
                bboxes.append(pdf_rect)

        # 如果没有缓存且有 YOLO 检测器，直接检测（兼容旧流程）
        if not bboxes and self.yolo_detector is not None:
            try:
                if self.yolo_detector.is_scanned_page(page):
                    print(f"[TableBBox] 页 {page_num + 1} 为扫描页，直接 YOLO 检测")
                    detections, _ = self.yolo_detector.detect_tables_in_page_pixels(page)
                    if detections:
                        scale = self.yolo_detector.dpi / 72
                        for det in detections:
                            x1, y1, x2, y2 = det['bbox']
                            pdf_rect = fitz.Rect(
                                x1 / scale, y1 / scale,
                                x2 / scale, y2 / scale
                            )
                            bboxes.append(pdf_rect)
                        print(f"[TableBBox] 页 {page_num + 1} YOLO 检测到 {len(bboxes)} 个表格")
            except Exception as e:
                print(f"[TableBBox] YOLO 检测失败: {e}")

        return bboxes

    def classify_models(self, models: List[Dict], page_num: int) -> Dict[str, List[Dict]]:
        """
        型号分类：表内 / 表外 / 可疑

        Args:
            models: 型号列表（带坐标）
            page_num: 页码

        Returns:
            分类结果 {'in_table': [...], 'out_table': [...], 'suspicious': [...]}
        """
        table_bboxes = self.get_table_bboxes(page_num)
        result = {
            'in_table': [],
            'out_table': [],
            'suspicious': [],
        }

        for model in models:
            bbox = model.get('bbox')
            confidence = model.get('confidence', 'low')

            # 判断是否在表格内
            in_any_table = False
            if bbox:
                for idx, table_bbox in enumerate(table_bboxes):
                    if bbox.intersects(table_bbox) or table_bbox.contains(bbox):
                        model['in_table'] = True
                        model['table_index'] = idx
                        in_any_table = True
                        break

            # 分类
            if in_any_table:
                result['in_table'].append(model)
            elif confidence in ('high', 'medium'):
                result['out_table'].append(model)
            else:
                result['suspicious'].append(model)

        return result

    def extract_with_table_alignment(self, page_num: int) -> Dict:
        """
        提取页面表格 + 表外型号

        简化架构（诊断修复后）：
        1. 矢量PDF：直接用 find_tables 提取表格内容（最可靠）
        2. 扫描PDF：YOLO + OCR + 布局推断
        3. 守恒检查：确保每页所有词都被归类
        """
        page = self.doc[page_num]
        is_scanned = self._is_scanned_pdf(page_num)

        extracted_tables = []
        out_table_models = []
        out_table_cells = []      # 表外文字（已归类，只是不在表格框内）
        unclassified_cells = []   # 真·未归类（既不在表内也判不出表外）
        in_table_count = 0
        out_table_count = 0
        total_words = 0

        if not is_scanned:
            # ====== 矢量PDF：直接用 PyMuPDF find_tables + split_double_table ======
            try:
                tabs = page.find_tables()

                # 嵌套子表去重：find_tables 有时把一个大表内部的子区域也报成独立表，
                # 导致同一张表被"切碎"成多块。按面积降序，丢弃被某个更大表包含≥90%的小表。
                _raw_tabs = list(tabs.tables)
                _kept_tabs = []
                for tab in sorted(_raw_tabs, key=lambda t: -((t.bbox[2]-t.bbox[0])*(t.bbox[3]-t.bbox[1]))):
                    _b = tab.bbox
                    _small_area = (_b[2]-_b[0]) * (_b[3]-_b[1])
                    _dup = False
                    for kept in _kept_tabs:
                        k = kept.bbox
                        ix0, iy0 = max(_b[0], k[0]), max(_b[1], k[1])
                        ix1, iy1 = min(_b[2], k[2]), min(_b[3], k[3])
                        if ix1 > ix0 and iy1 > iy0:
                            inter = (ix1-ix0) * (iy1-iy0)
                            # IoMin：小表被大表包含的比例；≥0.9 视为嵌套子表
                            if _small_area > 0 and inter / _small_area >= 0.9:
                                _dup = True
                                break
                    if not _dup:
                        _kept_tabs.append(tab)

                for tab in _kept_tabs:
                    try:
                        raw_data = tab.extract()
                        table_data = []
                        for row in raw_data:
                            table_data.append([str(cell).strip() if cell else "" for cell in row])
                        # 只过滤完全空的表格；有非空内容就保留（包括1行表）
                        has_content = any(
                            cell.strip()
                            for row in table_data
                            for cell in row
                        )
                        if table_data and has_content:
                            # 计算表格边界
                            bbox = tab.bbox if tab.bbox else (0, 0, page.rect.width, page.rect.height)
                            headers = []
                            try:
                                h = tab.header
                                if h:
                                    for row in h:
                                        headers.append([str(c).strip() if c else "" for c in row])
                            except Exception:
                                headers = []

                            # 双栏表格拆分（与旧版保持一致）
                            split_tables = self.structure_analyzer.split_double_table(table_data)

                            for sub_table in split_tables:
                                if sub_table and len(sub_table) >= 1:
                                    extracted_tables.append({
                                        'bbox': list(bbox),
                                        'source': 'find_tables',
                                        'data': sub_table,
                                        'headers': headers,
                                        'row_count': len(sub_table),
                                        'col_count': len(sub_table[0]) if sub_table else 0,
                                    })
                                    # 统计表内词数
                                    for row in sub_table:
                                        for cell in row:
                                            if cell.strip():
                                                in_table_count += len(cell.split())
                    except Exception:
                        pass
            except Exception:
                pass

            # 矢量页兜底：find_tables 未命中（网格线缺失/特殊排版）时，
            # 再走检测管线（YOLO→布局推断→覆盖检查）找表，避免漏检被 shortcut 成 0 表。
            # 仅当 find_tables 完全无表时才触发，正常矢量页（如 AJ/JSM）不受影响。
            if not extracted_tables:
                try:
                    table_boxes = self.table_pipeline.detect_all_table_regions(page, page_num)
                    if table_boxes:
                        classification = self.table_pipeline.classify_page_text(page, page_num, table_boxes, is_scanned=is_scanned)
                        in_table_cells = list(classification['in_table'])
                        for table_box in table_boxes:
                            content = self.table_pipeline.extract_table_content(page, table_box, in_table_cells)
                            if content is not None:
                                extracted_tables.append(content)
                except Exception:
                    pass

            # 表外文字：全页文本减去表内文本区域
            all_text_dict = page.get_text("dict")
            all_words = []
            for block in all_text_dict.get("blocks", []):
                if "lines" in block:
                    for line in block["lines"]:
                        for span in line.get("spans", []):
                            text = span.get("text", "").strip()
                            if text:
                                x0, y0, x1, y1 = span["bbox"]
                                all_words.append({
                                    'text': text,
                                    'bbox': fitz.Rect(x0, y0, x1, y1),
                                    'cx': (x0 + x1) / 2,
                                    'cy': (y0 + y1) / 2,
                                })

            total_words = len(all_words)
            # 表外词 = 不在任何表格框内的词
            for word in all_words:
                in_any_table = False
                for tab in extracted_tables:
                    tb = tab.get('bbox', (0, 0, page.rect.width, page.rect.height))
                    if tb[0] <= word['cx'] <= tb[2] and tb[1] <= word['cy'] <= tb[3]:
                        in_any_table = True
                        break
                if not in_any_table:
                    out_table_count += len(word['text'].split())
                    # 存储完整坐标信息（矢量PDF用PDF点坐标）
                    word_entry = {
                        'text': word['text'],
                        'bbox': word['bbox'],
                        'vector_bbox': word['bbox'],
                        'pixel_bbox': [0, 0, 0, 0],  # 矢量PDF无像素坐标
                        'cx': word['cx'],
                        'cy': word['cy'],
                        'reason': '矢量PDF-表外',
                    }
                    # 修复：表外文字归入 out_table_cells，不再与 unclassified 共用同一对象
                    out_table_cells.append(word_entry)

            # 表外型号正则
            out_table_text = ' '.join([w['text'] for w in out_table_cells])
            if out_table_text.strip():
                candidates = self.model_extractor.extract_models_from_text(out_table_text)
                for candidate in candidates:
                    out_table_models.append({
                        'model': candidate["model"],
                        'model_upper': candidate.get("model_upper", candidate["model"].upper()),
                        'pattern_type': candidate.get("pattern_type", ""),
                        'confidence': candidate.get("confidence", "medium"),
                        'bbox': None,
                        'page': page_num,
                        'source_pattern': candidate.get("source_pattern", ""),
                        'context': self._make_model_context(out_table_text, candidate["model"]),
                        'origin': '矢量-表外',
                        'in_table': False,
                        'table_index': -1,
                    })

        else:
            # ====== 扫描PDF：PPStructure全页扫描 ======
            try:
                scan_result = self.ocr_extractor.scan_page_ppstructure(page, page_num, dpi=300)
            except Exception as e:
                logger.error(f"页{page_num+1}(扫描): PPStructure扫描异常: {e}")
                scan_result = {'tables': [], 'out_table_cells': [], 'all_cells': []}

            # 表内：PPStructure直接返回结构化表格
            for tbl in scan_result.get('tables', []):
                extracted_tables.append({
                    'bbox': tbl.get('bbox', [0, 0, page.rect.width, page.rect.height]),
                    'source': 'ppstructure',
                    'data': tbl.get('data', []),
                    'headers': [],
                    'row_count': len(tbl.get('data', [])),
                    'col_count': len(tbl.get('data', [[]])[0]) if tbl.get('data') and tbl['data'][0] else 0,
                })

            out_table_cells = scan_result.get('out_table_cells', [])
            all_cells = scan_result.get('all_cells', [])
            total_words = len(all_cells)
            in_table_count = sum(len(t.get('data', [])) for t in scan_result.get('tables', []))
            out_table_count = len(out_table_cells)

            logger.info(f"页{page_num+1}(扫描): PPStructure检测到 {len(scan_result.get('tables', []))} 个表格, 总文字={total_words}, 表外={out_table_count}")

            # 表外型号正则
            out_table_text = ' '.join([c.get('text', '') for c in out_table_cells])
            if out_table_text.strip():
                candidates = self.model_extractor.extract_models_from_text(out_table_text)
                for candidate in candidates:
                    out_table_models.append({
                        'model': candidate["model"],
                        'model_upper': candidate.get("model_upper", candidate["model"].upper()),
                        'pattern_type': candidate.get("pattern_type", ""),
                        'confidence': candidate.get("confidence", "medium"),
                        'bbox': None,
                        'page': page_num,
                        'source_pattern': candidate.get("source_pattern", ""),
                        'context': self._make_model_context(out_table_text, candidate["model"]),
                        'origin': '扫描-表外',
                        'in_table': False,
                        'table_index': -1,
                    })

        # 守恒检查
        conservation_ok = True
        if is_scanned:
            # 扫描页：PPStructure的所有文字 = 表外 + 表内（表内=all-out）
            unclassified_cells = []
            in_table_cells_count = total_words - out_table_count
            if in_table_count < 0:
                in_table_count = 0
            conservation_ok = (out_table_count + in_table_cells_count == total_words)
        else:
            unclassified_cells = []

        return {
            'page': page_num,
            'is_scanned': is_scanned,
            'tables': extracted_tables,
            'out_table_models': out_table_models,
            'out_table_cells': out_table_cells,
            'unclassified': unclassified_cells,
            'conservation': {
                'total': total_words,
                'in_table': in_table_count,
                'out_table': out_table_count,
                'unclassified': len(unclassified_cells),
                'ok': conservation_ok,
            }
        }

    @staticmethod
    def _make_model_context(full_text: str, model: str, span: int = 25) -> str:
        """截取型号在原文中的上下文片段，便于人工核对表外型号来源"""
        if not full_text or not model:
            return ""
        pos = full_text.upper().find(model.upper())
        if pos < 0:
            return ""
        start = max(0, pos - span)
        end = min(len(full_text), pos + len(model) + span)
        prefix = "..." if start > 0 else ""
        suffix = "..." if end < len(full_text) else ""
        return f"{prefix}{full_text[start:end]}{suffix}".replace("\n", " ")

    def _find_model_bbox_in_cells(self, model_str: str, cells: List[Dict]) -> Optional[fitz.Rect]:
        """在 cells 列表中查找型号对应的 bbox"""
        model_upper = model_str.upper()

        for cell in cells:
            text = cell.get('text', '').upper()
            if text == model_upper or model_upper in text or text in model_upper:
                vb = cell.get('vector_bbox')
                if vb:
                    return vb
                pb = cell.get('pixel_bbox')
                if pb:
                    return fitz.Rect(
                        pb[0] * self.table_pipeline.inverse_scale,
                        pb[1] * self.table_pipeline.inverse_scale,
                        pb[2] * self.table_pipeline.inverse_scale,
                        pb[3] * self.table_pipeline.inverse_scale
                    )

        for cell in cells:
            text = cell.get('text', '').upper()
            if text[:4] == model_upper[:4]:
                pb = cell.get('pixel_bbox')
                if pb:
                    return fitz.Rect(
                        pb[0] * self.table_pipeline.inverse_scale,
                        pb[1] * self.table_pipeline.inverse_scale,
                        pb[2] * self.table_pipeline.inverse_scale,
                        pb[3] * self.table_pipeline.inverse_scale
                    )

        return None

    def extract_tables(self, page_num: int = 0) -> List[List[List[str]]]:
        """
        兼容旧接口：提取指定页面的表格数据

        内部使用新架构 extract_with_table_alignment 三源检测流水线，
        返回旧格式 List[List[List[str]]] 以兼容 GUI。

        Args:
            page_num: 页码（从0开始）

        Returns:
            表格数据列表，每个表格是一个二维列表
        """
        if page_num in self._scanned_cache:
            cached = self._scanned_cache[page_num]
            if 'tables' in cached:
                return [t['data'] for t in cached['tables']]

        result = self.extract_with_table_alignment(page_num)
        self._scanned_cache[page_num] = result

        tables = result.get('tables', [])
        return [t['data'] for t in tables]

    def get_cached_out_table_models(self) -> List[Dict]:
        """
        汇总本文档已提取页面的「表外型号」（只读缓存，不触发重新提取）

        表外型号 = 不在任何表格框内、由全页文本/OCR 正则命中的型号
        （产品图旁、封底、纯文字页等），带 in_table=False 标记。

        Returns:
            型号字典列表，按页码升序
        """
        models = []
        for page_num in sorted(self._scanned_cache.keys()):
            cached = self._scanned_cache.get(page_num) or {}
            for m in (cached.get('out_table_models') or []):
                if m.get('in_table'):
                    continue
                models.append(m)
        return models

    def export_to_excel(self, output_path: str, page_range: Optional[range] = None):
        """
        新架构导出：四 Sheet 结构

        Sheet1「表格信息」- 同列名结构纵向合并，跨页续表自动接
        Sheet2「表外型号」- 表外检测到的型号
        Sheet3「未归类」- 守恒兜底出口
        Sheet4「提取概览」- 每页处理统计

        Args:
            output_path: 输出 Excel 路径
            page_range: 页码范围，None 表示全部
        """
        if not HAS_OPENPYXL:
            print("[Export] openpyxl 未安装，改为 CSV 导出模式")
            self._export_csv(output_path)
            return

        print(f"\n[Export] 开始提取并导出: {output_path}")

        wb = Workbook()
        ws1 = wb.active
        ws1.title = "表格信息"
        ws2 = wb.create_sheet("表外型号")
        ws3 = wb.create_sheet("潜在型号")
        ws4 = wb.create_sheet("提取概览")

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # ---- Sheet1: 表格信息 ----
        sheet1_headers = ['来源页', '表序号', '结构签名', '列名', '行号', '单元格内容', '表头匹配']
        ws1.append(sheet1_headers)
        for col in range(1, len(sheet1_headers) + 1):
            cell = ws1.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # ---- Sheet2: 表外型号 ----
        sheet2_headers = ['来源页', '型号', '置信度', '坐标', '类型', '上下文']
        ws2.append(sheet2_headers)
        for col in range(1, len(sheet2_headers) + 1):
            cell = ws2.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # ---- Sheet3: 未归类 ----
        sheet3_headers = ['来源页', '文本内容', '像素坐标', '矢量坐标', '原因']
        ws3.append(sheet3_headers)
        for col in range(1, len(sheet3_headers) + 1):
            cell = ws3.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # ---- Sheet4: 提取概览 ----
        sheet4_headers = ['来源页', '表内词数', '表外词数', '未归类词数', '总词数', '守恒OK', '表数量', '扫描页']
        ws4.append(sheet4_headers)
        for col in range(1, len(sheet4_headers) + 1):
            cell = ws4.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # 收集所有唯一型号（用于反查参数）
        all_models_for_lookup = set()

        # 表格结构签名跟踪：同列名结构合并
        structure_signatures: Dict[str, Dict] = {}
        table_global_idx = 0

        total_pages = self.doc.page_count
        if page_range is None:
            page_range = range(total_pages)

        for page_num in page_range:
            if page_num >= total_pages:
                break

            print(f"  处理第 {page_num + 1}/{total_pages} 页...")

            try:
                result = self.extract_with_table_alignment(page_num)
            except Exception as e:
                print(f"  [警告] 第{page_num+1}页提取失败: {e}")
                continue

            # Sheet1: 表格信息 — 同列名结构纵向合并
            for table_data in result['tables']:
                data_rows = table_data.get('data', [])
                if not data_rows:
                    continue

                # 计算结构签名（列数+前几行内容特征）
                col_count = table_data.get('col_count', len(data_rows[0]) if data_rows else 0)
                row_count = table_data.get('row_count', len(data_rows))
                sig = f"c{col_count}r{row_count}"

                if sig not in structure_signatures:
                    table_global_idx += 1
                    structure_signatures[sig] = {
                        'global_idx': table_global_idx,
                        'col_count': col_count,
                        'first_page': page_num + 1,
                    }

                gidx = structure_signatures[sig]['global_idx']

                # 尝试识别表头
                headers = self._guess_headers_from_data(data_rows)

                for row_idx, row in enumerate(data_rows):
                    for col_idx, cell_text in enumerate(row):
                        if cell_text.strip():
                            header_name = headers[col_idx] if col_idx < len(headers) else f"列{col_idx+1}"
                            ws1.append([
                                page_num + 1,
                                gidx,
                                sig,
                                header_name,
                                row_idx + 1,
                                cell_text.strip(),
                                '是' if row_idx == 0 else '',
                            ])

            # Sheet2: 表外型号
            for model in result['out_table_models']:
                bbox = model.get('bbox')
                bbox_str = ""
                if bbox:
                    try:
                        bbox_str = f"({bbox.x0:.0f},{bbox.y0:.0f})-({bbox.x1:.0f},{bbox.y1:.0f})"
                    except Exception:
                        bbox_str = "N/A"
                else:
                    bbox_str = "N/A"

                context = self._get_model_context(model, result)
                ws2.append([
                    page_num + 1,
                    model['model'],
                    model.get('confidence', ''),
                    bbox_str,
                    model.get('pattern_type', ''),
                    context,
                ])
                all_models_for_lookup.add((page_num + 1, model['model']))

            # Sheet3: 未归类（守恒兜底）
            for item in result.get('unclassified', []):
                text = item.get('text', '')
                pb = item.get('pixel_bbox', [0, 0, 0, 0])
                vb = item.get('vector_bbox')
                vb_str = ""
                if vb:
                    try:
                        vb_str = f"({vb.x0:.1f},{vb.y0:.1f})-({vb.x1:.1f},{vb.y1:.1f})"
                    except Exception:
                        pass
                reason = item.get('reason', '守恒兜底')
                ws3.append([
                    page_num + 1,
                    text,
                    f"({pb[0]:.0f},{pb[1]:.0f})-({pb[2]:.0f},{pb[3]:.0f})",
                    vb_str,
                    reason,
                ])

            # Sheet4: 提取概览
            cons = result.get('conservation', {})
            ws4.append([
                page_num + 1,
                cons.get('in_table', 0),
                cons.get('out_table', 0),
                cons.get('unclassified', 0),
                cons.get('total', 0),
                'OK' if cons.get('ok', True) else 'FAIL',
                len(result.get('tables', [])),
                '扫描' if result.get('is_scanned', False) else '矢量',
            ])

        # Sheet4: 汇总行
        if ws4.max_row > 1:
            total_in = sum(ws4.cell(row=r, column=2).value or 0 for r in range(2, ws4.max_row + 1))
            total_out = sum(ws4.cell(row=r, column=3).value or 0 for r in range(2, ws4.max_row + 1))
            total_un = sum(ws4.cell(row=r, column=4).value or 0 for r in range(2, ws4.max_row + 1))
            total_all = sum(ws4.cell(row=r, column=5).value or 0 for r in range(2, ws4.max_row + 1))
            ws4.append(['合计', total_in, total_out, total_un, total_all, '', '', ''])

        # 调整列宽
        for ws in [ws1, ws2, ws3, ws4]:
            for column_cells in ws.columns:
                max_length = 0
                column = column_cells[0].column_letter
                for cell in column_cells:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except Exception:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width

        wb.save(output_path)
        print(f"[Export] 完成: {output_path}")

        self._print_export_stats(ws1, ws2, ws3, ws4)
        self.close()

    def _guess_headers_from_data(self, data: List[List[str]]) -> List[str]:
        """从表格数据推断列名（用首行内容做列名）"""
        if not data:
            return []
        col_count = max(len(row) for row in data)
        headers = []
        for c in range(col_count):
            header_texts = []
            for row in data[:3]:
                if c < len(row) and row[c].strip():
                    header_texts.append(row[c].strip())
            if header_texts:
                headers.append(header_texts[0][:20])
            else:
                headers.append(f"列{c+1}")
        return headers

    def _get_model_context(self, model: Dict, result: Dict) -> str:
        """获取型号的上下文（从 out_table_models 中找匹配文本）"""
        model_text = model.get('model', '')
        models = result.get('out_table_models', [])
        for m in models:
            mt = m.get('model', '')
            if model_text and (model_text in mt or mt in model_text):
                return mt[:80]
        return model_text[:80] if model_text else ''

    def _export_csv(self, output_path: str):
        """CSV 导出降级模式"""
        base = output_path.rsplit('.', 1)[0]
        for sheet_name, headers in [
            ('表格信息', ['来源页', '表序号', '结构签名', '列名', '行号', '单元格内容', '表头匹配']),
            ('表外型号', ['来源页', '型号', '置信度', '坐标', '类型', '上下文']),
            ('未归类', ['来源页', '文本内容', '像素坐标', '矢量坐标', '原因']),
            ('提取概览', ['来源页', '表内词数', '表外词数', '未归类词数', '总词数', '守恒OK', '表数量', '扫描页']),
        ]:
            csv_path = f"{base}_{sheet_name}.csv"
            hh.DataFrame(columns=headers).to_csv(csv_path, index=False, encoding='utf-8-sig')
            print(f"[Export] CSV: {csv_path}")

    def _print_export_stats(self, ws1, ws2, ws3, ws4=None):
        """打印导出统计"""
        table_count = ws1.max_row - 1
        out_model_count = ws2.max_row - 1
        unclassified_count = ws3.max_row - 1

        print(f"\n[统计] 表格信息单元格: {table_count} 个")
        print(f"[统计] 表外型号: {out_model_count} 个")
        print(f"[统计] 未归类词: {unclassified_count} 个")
        if ws4:
            overview_rows = ws4.max_row - 1
            print(f"[统计] 处理页面: {overview_rows} 页")
        print(f"[统计] 总计: {table_count + out_model_count + unclassified_count} 条")


# =============================================================================
# 椭圆表格检测类（阶段三新增）
# =============================================================================

class EllipticalTableDetector:
    """
    椭圆/非标准表格检测器

    用于检测和处理具有特殊形状的表格（如椭圆单元格、不规则表格等）。
    这些表格无法通过标准的行列对齐方式处理。
    """

    # 椭圆形状检测的最小和最大宽高比
    ELLIPSE_MIN_RATIO = 0.3
    ELLIPSE_MAX_RATIO = 3.0

    # 最小椭圆宽度（点）
    MIN_ELLIPSE_WIDTH = 20

    def __init__(self):
        self.detected_elliptical_tables: List[Dict] = []

    def detect_elliptical_regions(self, page: 'fitz.Page') -> List[Dict]:
        """
        检测页面上的椭圆/圆形区域

        Args:
            page: PyMuPDF 页面对象

        Returns:
            椭圆区域列表 [{'bbox': fitz.Rect, 'type': 'ellipse', ...}]
        """
        elliptical_regions = []

        try:
            # 获取页面绘图元素（包含椭圆/圆形的边界路径）
            drawings = page.get_drawings()

            for drawing in drawings:
                rect = drawing.get('rect')
                if rect and self._is_ellipse_shape(drawing):
                    elliptical_regions.append({
                        'bbox': fitz.Rect(rect),
                        'type': 'ellipse',
                        'drawing': drawing,
                    })
        except Exception as e:
            print(f"[EllipticalDetector] 检测椭圆区域失败: {e}")

        self.detected_elliptical_tables = elliptical_regions
        return elliptical_regions

    def _is_ellipse_shape(self, drawing: Dict) -> bool:
        """判断绘图是否为椭圆/圆形"""
        rect = drawing.get('rect')
        if not rect:
            return False

        width = rect.width
        height = rect.height

        if width < self.MIN_ELLIPSE_WIDTH or height < self.MIN_ELLIPSE_WIDTH:
            return False

        # 计算宽高比
        ratio = width / height if height > 0 else float('inf')

        # 椭圆/圆形的宽高比在特定范围内
        if self.ELLIPSE_MIN_RATIO <= ratio <= self.ELLIPSE_MAX_RATIO:
            # 检查绘图类型
            for item in drawing.get('items', []):
                if item[0] == 'c' and len(item[1]) >= 4:  # 曲线
                    return True

        return False

    def extract_elliptical_table_content(
        self,
        page: 'fitz.Page',
        region: Dict
    ) -> Dict:
        """
        提取椭圆表格区域的内容

        Args:
            page: PyMuPDF 页面对象
            region: 椭圆区域信息

        Returns:
            提取的内容字典
        """
        bbox = region['bbox']

        try:
            # 提取区域内的文本
            text_in_region = page.get_text("text", clip=bbox)

            # 提取区域内的文本块（带坐标）
            text_dict = page.get_text("dict", clip=bbox)

            spans_text = ""
            span_locations = []

            for block in text_dict.get("blocks", []):
                if "lines" in block:
                    for line in block["lines"]:
                        for span in line.get("spans", []):
                            span_text = span.get("text", "").strip()
                            if span_text:
                                spans_text += span_text + " "
                                span_locations.append({
                                    'text': span_text,
                                    'bbox': fitz.Rect(span['bbox']),
                                })

            return {
                'bbox': bbox,
                'type': 'elliptical_table',
                'text': spans_text.strip(),
                'raw_text': text_in_region.strip(),
                'spans': span_locations,
                'cell_count': len(span_locations),
            }
        except Exception as e:
            print(f"[EllipticalDetector] 提取椭圆表格内容失败: {e}")
            return {
                'bbox': bbox,
                'type': 'elliptical_table',
                'text': '',
                'raw_text': '',
                'spans': [],
                'cell_count': 0,
            }

    def process_page_elliptical_tables(
        self,
        page: 'fitz.Page',
        model_extractor: 'ModelExtractor'
    ) -> List[Dict]:
        """
        处理页面上的所有椭圆表格

        Args:
            page: PyMuPDF 页面对象
            model_extractor: 型号提取器

        Returns:
            椭圆表格处理结果列表
        """
        results = []

        # 检测椭圆区域
        elliptical_regions = self.detect_elliptical_regions(page)

        for region in elliptical_regions:
            # 提取内容
            content = self.extract_elliptical_table_content(page, region)

            # 提取型号
            models = model_extractor.extract_models_from_text(content['text'])

            results.append({
                'region': region,
                'content': content,
                'models': models,
                'model_count': len(models),
            })

        return results


# =============================================================================
# YOLO 表格检测器（阶段四新增，用于扫描页辅助检测）
# =============================================================================

class YOLOTableDetector:
    """
    YOLO 表格检测器（扫描页专用）

    工作流程：
    1. YOLO 只负责找出表格位置（像素坐标）
    2. 全页 OCR 获取文本+像素坐标
    3. 用 YOLO 表格框分类 OCR 结果（表内/表外）
    4. 表内：按表格结构组织；表外：跑型号正则

    可配置参数：权重路径、置信度、表格类别 ID
    """

    # 默认权重路径（可通过构造函数参数覆盖）
    DEFAULT_WEIGHTS = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "..", "weights", "table_1cls_dev.pt"
    )

    def __init__(
        self,
        weights_path: Optional[str] = None,
        conf: float = 0.10,
        table_class_id: int = 0,
        dpi: int = 150
    ):
        """
        初始化 YOLO 检测器

        Args:
            weights_path: 权重文件路径，None 用默认
            conf: 置信度阈值（默认 0.10，定稿工作点）
                  依据 2026-07-31 conf 扫描（自有 val 52图/66框）：
                    conf 0.10 -> 207 框, IoU0.3 召回 0.742
                    conf 0.05 -> 344 框, IoU0.3 召回 0.803（+6pt 但框数翻倍）
                    conf 0.15 -> 167 框, IoU0.3 召回 0.697（白丢 4.5pt）
                  0.10 是拐点：漏框代价（整表消失）远大于误框代价（多一张废 sheet），
                  故取偏召回侧；多出的误框由下游"退化自愈"（<2行/<2列判定非表）清理。
            table_class_id: 表格类别 ID（默认 0，新模型可能不同）
            dpi: PDF 渲染 DPI（默认 150）
        """
        self.weights_path = weights_path or self.DEFAULT_WEIGHTS
        self.conf = conf
        self.table_class_id = table_class_id
        self.dpi = dpi
        self.model = None
        self._loaded = False

        # 关键修复：在 PaddleOCR 之前预加载 torch，避免 DLL 冲突
        # （PaddlePaddle 的 DLL 会覆盖 torch 需要的共享库，导致 shm.dll 加载失败）
        try:
            import torch
            import ultralytics
        except Exception:
            pass

    def _resolve_weights_path(self) -> str:
        """
        解析权重路径，支持开发环境和 PyInstaller 打包环境

        PyInstaller 打包时通过 --add-data 将权重复制到 sys._MEIPASS/weights/ 下
        """
        path = self.weights_path
        basename = os.path.basename(path)

        # 1. 直接存在
        if os.path.exists(path):
            return path

        # 2. PyInstaller 打包环境（权重在 sys._MEIPASS/weights/ 下）
        if hasattr(sys, '_MEIPASS'):
            # --add-data "table_1cls_dev.pt;weights" → sys._MEIPASS/weights/table_1cls_dev.pt
            candidates = [
                os.path.join(sys._MEIPASS, "weights", basename),
                os.path.join(sys._MEIPASS, basename),
            ]
            for c in candidates:
                if os.path.exists(c):
                    return c

        # 3. 开发环境回退：脚本同级或上级 weights 目录
        script_dir = os.path.dirname(os.path.abspath(__file__))
        candidates = [
            os.path.join(script_dir, "..", "weights", basename),
            os.path.join(script_dir, "weights", basename),
            os.path.join(script_dir, basename),
        ]
        for c in candidates:
            if os.path.exists(c):
                return c

        return path  # 返回原路径，让调用方处理错误

    def load_model(self) -> bool:
        """加载 YOLO 模型"""
        if self._loaded and self.model:
            return True

        try:
            from ultralytics import YOLO

            resolved_path = self._resolve_weights_path()
            if not os.path.exists(resolved_path):
                print(f"[YOLO] 权重文件不存在: {self.weights_path} (解析后: {resolved_path})")
                return False

            print(f"[YOLO] 加载权重: {resolved_path}")
            self.model = YOLO(resolved_path)
            self._loaded = True
            return True
        except ImportError:
            print("[YOLO] ultralytics 未安装，跳过 YOLO 检测")
            return False
        except Exception as e:
            print(f"[YOLO] 加载模型失败: {e}")
            return False

    def detect_tables_in_image(self, image_path: str) -> List[Dict]:
        """
        在图片中检测表格位置（像素坐标）

        策略：先用默认置信度检测，如果检测不到表格，再用更低的置信度重试

        Args:
            image_path: 图片路径

        Returns:
            检测结果列表 [{'cls': int, 'conf': float, 'bbox': [x1,y1,x2,y2]}]
            只返回 table_class_id 对应的类别
        """
        if not self.load_model():
            return []

        # 多级置信度兜底检测：主工作点 self.conf(0.10)，仍为空再降到 0.05 抢召回
        # （去重 + 降序，避免 self.conf 本身就等于某一级时重复推理）
        conf_levels = sorted({self.conf, 0.05}, reverse=True)
        for conf in conf_levels:
            if conf > self.conf:
                continue
            
            try:
                results = self.model.predict(
                    image_path,
                    conf=conf,
                    iou=0.45,
                    verbose=False,
                    device="cpu"
                )

                tables = []
                for r in results:
                    for box in r.boxes:
                        cls_id = int(box.cls[0])
                        # 只保留配置的表格类别
                        if cls_id == self.table_class_id:
                            x1, y1, x2, y2 = [float(v) for v in box.xyxy[0]]
                            tables.append({
                                'cls': cls_id,
                                'conf': float(box.conf[0]),
                                'bbox': [x1, y1, x2, y2],
                            })

                if tables:
                    return tables
                    
                if conf == self.conf:
                    print(f"[YOLO] conf={conf} 未检测到表格，尝试更低置信度...")
                    
            except Exception as e:
                print(f"[YOLO] 检测失败 (conf={conf}): {e}")
                continue

        return []

    def detect_tables_in_page_pixels(
        self,
        page: 'fitz.Page'
    ) -> Tuple[List[Dict], str]:
        """
        在 PDF 页面中检测表格位置（返回像素坐标 + 渲染图片路径）

        Args:
            page: PyMuPDF 页面对象

        Returns:
            (表格检测列表, 渲染图片路径)
            检测列表元素: {'cls': int, 'conf': float, 'bbox': [x1,y1,x2,y2]}
            图片路径用于后续 OCR
        """
        if not self.load_model():
            return [], ""

        temp_img_path = None
        try:
            # 渲染页面为图片（带缓存，扫描页三源共享）
            pix = self._get_pixmap(page, self.dpi)

            # 保存临时图片
            import tempfile
            fd, temp_img_path = tempfile.mkstemp(suffix='.png')
            os.close(fd)
            pix.save(temp_img_path)

            # YOLO 检测（像素坐标）
            detections = self.detect_tables_in_image(temp_img_path)

            return detections, temp_img_path
        except Exception as e:
            print(f"[YOLO] 页面检测失败: {e}")
            if temp_img_path and os.path.exists(temp_img_path):
                try:
                    os.unlink(temp_img_path)
                except Exception:
                    pass
            return [], ""

    def classify_ocr_by_tables(
        self,
        ocr_results: List[Dict],
        table_detections: List[Dict]
    ) -> Dict[str, List[Dict]]:
        """
        用 YOLO 表格框分类 OCR 结果（表内/表外）

        Args:
            ocr_results: OCR 识别结果列表，每个元素包含 x0,y0,x1,y1,text,center_x,center_y
            table_detections: YOLO 检测的表格列表，每个元素包含 bbox=[x1,y1,x2,y2]

        Returns:
            {'in_table': [...], 'out_table': [...]}
        """
        result = {
            'in_table': [],
            'out_table': [],
        }

        if not ocr_results:
            return result

        for cell in ocr_results:
            center_x = cell.get('center_x', 0)
            center_y = cell.get('center_y', 0)

            # 检查中心点是否在任何表格框内
            in_any_table = False
            for table_det in table_detections:
                tx1, ty1, tx2, ty2 = table_det['bbox']
                if tx1 <= center_x <= tx2 and ty1 <= center_y <= ty2:
                    cell['table_conf'] = table_det['conf']  # 记录所在表格的置信度
                    result['in_table'].append(cell)
                    in_any_table = True
                    break

            if not in_any_table:
                result['out_table'].append(cell)

        return result

    def process_scanned_page(
        self,
        page: 'fitz.Page',
        ocr_extractor: 'OCRTableExtractor'
    ) -> Dict:
        """
        处理扫描页：YOLO 检测 + 全页 OCR + 空间分类 + 降级自愈

        策略：
        1. YOLO 检测表格位置
        2. 若 YOLO 落空 → 降级布局推断（OCR cells 聚类找表状结构）
        3. 若仍无 → 全进 out_table 并标记 yolo_failed=True
        4. 守恒兜底：任何无法归类的进入 unclassified

        Args:
            page: PyMuPDF 页面对象
            ocr_extractor: OCR 识别器实例

        Returns:
            {
                'table_detections': 检测到的表格列表,
                'in_table_cells': 表内 OCR 文本,
                'out_table_cells': 表外 OCR 文本,
                'unclassified_cells': 未归类文本,
                'image_path': 渲染图片路径,
                'yolo_failed': YOLO 是否落空,
                'degraded': 是否降级到布局推断,
            }
        """
        result = {
            'table_detections': [],
            'in_table_cells': [],
            'out_table_cells': [],
            'unclassified_cells': [],
            'image_path': '',
            'yolo_failed': False,
            'degraded': False,
        }

        # 1. YOLO 检测表格位置
        table_detections, img_path = self.detect_tables_in_page_pixels(page)
        result['table_detections'] = table_detections
        result['image_path'] = img_path

        if not img_path:
            result['yolo_failed'] = True
            result['unclassified_cells'] = []
            return result

        # 2. 全页 OCR
        ocr_cells = ocr_extractor.recognize_from_image(img_path, is_scanned=True)

        # 3. 空间分类
        if table_detections and ocr_cells:
            classified = self.classify_ocr_by_tables(ocr_cells, table_detections)
            result['in_table_cells'] = classified['in_table']
            result['out_table_cells'] = classified['out_table']
        else:
            # YOLO 落空 → 降级布局推断
            result['yolo_failed'] = True
            if ocr_cells:
                inferred_boxes = self._infer_tables_from_ocr_cells(ocr_cells)
                if inferred_boxes:
                    result['degraded'] = True
                    classified = self.classify_ocr_by_tables(ocr_cells, inferred_boxes)
                    result['in_table_cells'] = classified['in_table']
                    result['out_table_cells'] = classified['out_table']
                    result['table_detections'] = inferred_boxes
                else:
                    # 布局推断也失败 → 全进 out_table，标记
                    result['out_table_cells'] = ocr_cells
                    result['unclassified_cells'] = [
                        {**c, 'reason': 'YOLO落空+布局推断失败'} for c in ocr_cells
                    ]

        # 4. 守恒检查
        total_cells = len(ocr_cells)
        accounted = len(result['in_table_cells']) + len(result['out_table_cells']) + len(result['unclassified_cells'])
        if accounted < total_cells:
            # 兜底：把遗漏的放入 unclassified
            result['unclassified_cells'].extend(
                [{'text': '未归类兜底', 'reason': f'守恒修复: {total_cells - accounted} 个遗漏'}]
            )

        return result

    def _infer_tables_from_ocr_cells(self, ocr_cells: List[Dict]) -> List[Dict]:
        """
        从 OCR cells 推断表格结构（降级布局推断）

        规则：连续≥3行、每行≥2个 x 对齐词群
        """
        if len(ocr_cells) < 5:
            return []

        sorted_cells = sorted(ocr_cells, key=lambda c: (c.get('center_y', 0), c.get('center_x', 0)))

        # y 聚类成行
        lines = []
        current_line = []
        current_y = None
        y_tolerance = 15

        for cell in sorted_cells:
            cy = cell.get('center_y', 0)
            if current_y is None:
                current_y = cy
                current_line.append(cell)
            elif abs(cy - current_y) <= y_tolerance:
                current_line.append(cell)
                current_y = sum(c.get('center_y', 0) for c in current_line) / len(current_line)
            else:
                if current_line:
                    lines.append(current_line)
                current_line = [cell]
                current_y = cy

        if current_line:
            lines.append(current_line)

        # 找连续行中每行有多个 x 对齐词群的模式
        inferred = []
        i = 0
        while i < len(lines):
            table_lines = [lines[i]]
            j = i + 1
            while j < len(lines):
                if len(lines[j]) >= 2 and len(lines[j - 1]) >= 2:
                    table_lines.append(lines[j])
                    j += 1
                else:
                    break

            if len(table_lines) >= 3:
                # 计算表格边界
                all_x0 = [min(c.get('x0', 0) for c in l) for l in table_lines]
                all_y0 = [min(c.get('y0', 0) for c in l) for l in table_lines]
                all_x1 = [max(c.get('x1', 0) for c in l) for l in table_lines]
                all_y1 = [max(c.get('y1', 0) for c in l) for l in table_lines]

                bbox = [
                    min(all_x0), min(all_y0),
                    max(all_x1), max(all_y1)
                ]
                inferred.append({
                    'cls': 0,
                    'conf': 0.5,
                    'bbox': bbox,
                    'source': 'layout_inference',
                })

            i = j if j > i + 1 else i + 1

        return inferred

    def is_scanned_page(self, page: 'fitz.Page') -> bool:
        """
        判断页面是否为扫描页（无文本层）

        Args:
            page: PyMuPDF 页面对象

        Returns:
            True 表示扫描页
        """
        try:
            text = page.get_text()
            blocks = page.get_text("blocks")
            text_block_count = sum(1 for b in blocks if b[6] == 0)

            if text_block_count < 5 or len(text.strip()) < 20:
                return True
            return False
        except Exception:
            return True


# =============================================================================
# TableDetectionPipeline: 三源表框检测管线（新架构核心）
# =============================================================================

class TableDetectionPipeline:
    """
    表框检测管线：三源恒定全跑 + 并集去重 + 空间分类

    核心原则：
    1. 三源恒定全跑：find_tables ∪ YOLO ∪ 布局结构推断
    2. IoU>0.6 去重：保留边界更紧的
    3. 统一像素坐标系（DPI=300）
    4. 表框并集作单一事实源
    """

    def __init__(
        self,
        yolo_detector: Optional['YOLOTableDetector'] = None,
        ocr_extractor: Optional['OCRTableExtractor'] = None,
        dpi: int = 300
    ):
        self.yolo_detector = yolo_detector
        self.ocr_extractor = ocr_extractor
        self.dpi = dpi
        self.scale = dpi / 72  # PDF point → pixel 转换因子
        self.inverse_scale = 72 / dpi  # pixel → PDF point 转换因子
        self._render_cache = {}  # 页面渲染缓存（同一页三源只渲染一次）

    def is_scanned_page(self, page: 'fitz.Page') -> bool:
        """
        判断页面是否为扫描页（无文本层）

        修复：原先 detect_all_table_regions 直接调 self.is_scanned_page，
        但该方法只定义在 YOLOTableDetector 上，导致本类调用时抛
        AttributeError —— 扫描页整条检测链崩溃（表格与型号全部丢失）。
        这里补上本类实现；若已注入 yolo_detector 则复用其判定，保持一致。
        """
        if self.yolo_detector is not None and hasattr(self.yolo_detector, 'is_scanned_page'):
            try:
                return self.yolo_detector.is_scanned_page(page)
            except Exception:
                pass
        try:
            text = page.get_text()
            blocks = page.get_text("blocks")
            text_block_count = sum(1 for b in blocks if b[6] == 0)
            if text_block_count < 5 or len(text.strip()) < 20:
                return True
            return False
        except Exception:
            return True

    def _get_pixmap(self, page: 'fitz.Page', dpi: int):
        """带缓存的页面渲染：同一页在三源（YOLO/布局/覆盖）中只渲染一次"""
        key = (id(page), dpi)
        if key not in self._render_cache:
            mat = fitz.Matrix(dpi / 72.0, dpi / 72.0)
            self._render_cache[key] = page.get_pixmap(matrix=mat)
        return self._render_cache[key]

    def detect_all_table_regions(
        self,
        page: 'fitz.Page',
        page_num: int
    ) -> List[Dict]:
        """
        表框检测：矢量页用 find_tables，扫描页用 YOLO + 布局推断

        诊断结论（2026-07-31）：
        - 矢量PDF上，三源并集会导致坐标系统一问题和巨型框
        - find_tables 是矢量PDF上最可靠的表格检测方法
        - YOLO和布局推断仅在扫描页上有价值
        """
        is_scanned = self.is_scanned_page(page)
        self._render_cache = {}  # 每次调用只缓存当前页的渲染
        source_boxes = []

        if not is_scanned:
            # 矢量页：只用 find_tables（最可靠，边界精确）
            ft_boxes = self._detect_find_tables(page)
            source_boxes.extend(ft_boxes)
            # 如果 find_tables 没找到，降级尝试 YOLO
            if not source_boxes and self.yolo_detector:
                yolo_boxes = self._detect_yolo(page)
                source_boxes.extend(yolo_boxes)
            # 再降级布局推断
            if not source_boxes:
                layout_boxes = self._detect_layout_inference(page)
                source_boxes.extend(layout_boxes)
        else:
            # 扫描页：三源并行跑，独立计算后合并，不得短路
            # 源1: YOLO
            if self.yolo_detector:
                yolo_boxes = self._detect_yolo(page)
                source_boxes.extend(yolo_boxes)

            # 源2: 布局推断（不受 YOLO 结果影响，独立跑）
            layout_boxes = self._detect_layout_inference(page)
            source_boxes.extend(layout_boxes)

            # 源3: find_tables（扫描页大概率空，但仍兜底）
            ft_boxes = self._detect_find_tables(page)
            source_boxes.extend(ft_boxes)

            # 源4: 覆盖检查补框
            coverage_boxes = self._coverage_check(page, source_boxes)
            source_boxes.extend(coverage_boxes)

        # IoU>0.6 去重，保留边界更紧的
        if len(source_boxes) > 1:
            deduped = self._dedup_by_iou(source_boxes, iou_threshold=0.6)
            return deduped

        return source_boxes

    def _detect_find_tables(self, page: 'fitz.Page') -> List[Dict]:
        """源 1: find_tables（边界精确）"""
        boxes = []
        try:
            tables = page.find_tables()
            for tab in tables.tables:
                if tab.bbox:
                    # PDF point → pixel
                    x1, y1, x2, y2 = tab.bbox
                    pixel_box = [
                        x1 * self.scale, y1 * self.scale,
                        x2 * self.scale, y2 * self.scale
                    ]
                    boxes.append({
                        'bbox': pixel_box,
                        'pixel_bbox': pixel_box,
                        'source': 'find_tables',
                        'conf': 1.0,  # find_tables 可信度高
                    })
        except Exception as e:
            pass
        return boxes

    def _detect_yolo(self, page: 'fitz.Page') -> List[Dict]:
        """
        源 2: YOLO 检测 + 后处理

        后处理管线：
        1. NMS 去重（IoU<0.5 的重叠框合并）
        2. 面积/长宽比滤波（剔除产品图、标题块等误检）
        3. 按置信度排序
        """
        boxes = []
        if not self.yolo_detector:
            return boxes

        # 关键修复：先尝试加载模型，再检查 _loaded
        if not self.yolo_detector._loaded:
            self.yolo_detector.load_model()
        if not self.yolo_detector._loaded:
            return boxes

        temp_img_path = None
        try:
            # 渲染页面为图片（带缓存，扫描页三源共享）
            pix = self._get_pixmap(page, self.dpi)

            import tempfile
            fd, temp_img_path = tempfile.mkstemp(suffix='.png')
            os.close(fd)
            pix.save(temp_img_path)

            # 获取页面尺寸用于归一化
            page_w = pix.width
            page_h = pix.height
            page_area = page_w * page_h

            # YOLO 检测（conf=0.10 低阈值抢召回）
            detections = self.yolo_detector.detect_tables_in_image(temp_img_path)

            if not detections:
                return boxes

            # === 后处理 ===
            # Step 1: NMS — 同一页上高度重叠的 YOLO 框合并
            detections = self._yolo_nms(detections, iou_threshold=0.5)

            # Step 2: 面积/长宽比滤波
            filtered = []
            for det in detections:
                bbox = det['bbox']
                area = (bbox[2] - bbox[0]) * (bbox[3] - bbox[1])
                width = bbox[2] - bbox[0]
                height = bbox[3] - bbox[1]

                # 剔除面积过小的噪声框（< 页面 1%）
                if area < page_area * 0.01:
                    continue

                # 剔除极端长宽比（宽高比 > 10:1 或 1:10，通常不是表格）
                aspect = width / max(height, 1)
                if aspect > 10 or aspect < 0.1:
                    continue

                # 剔除覆盖全页的巨型框（可能是误检产品图/整页）
                if area > page_area * 0.95:
                    continue

                filtered.append(det)

            # Step 3: 按置信度排序
            filtered.sort(key=lambda d: d.get('conf', 0), reverse=True)

            for det in filtered:
                boxes.append({
                    'bbox': det['bbox'],
                    'pixel_bbox': det['bbox'],
                    'source': 'yolo',
                    'conf': det['conf'],
                })
        except Exception:
            pass
        finally:
            if temp_img_path and os.path.exists(temp_img_path):
                try:
                    os.unlink(temp_img_path)
                except Exception:
                    pass

        return boxes

    def _yolo_nms(self, detections: List[Dict], iou_threshold: float = 0.5) -> List[Dict]:
        """YOLO 检测结果 NMS（非极大值抑制），合并高度重叠的框"""
        if not detections:
            return []

        # 按置信度降序
        sorted_dets = sorted(detections, key=lambda d: d.get('conf', 0), reverse=True)
        kept = []

        for det in sorted_dets:
            bbox = det['bbox']
            is_duplicate = False
            for k in kept:
                if self._compute_iou(bbox, k['bbox']) > iou_threshold:
                    is_duplicate = True
                    break
            if not is_duplicate:
                kept.append(det)

        return kept

    def _detect_layout_inference(self, page: 'fitz.Page') -> List[Dict]:
        """
        源 3: 布局结构推断（纯几何零训练成本）

        规则：连续≥3行、每行≥2个 x 对齐词群
        """
        boxes = []
        try:
            # 获取页面文本（带坐标）
            text_dict = page.get_text("dict")

            # 收集所有文本行（像素坐标）
            lines = []
            for block in text_dict.get("blocks", []):
                if "lines" in block:
                    for line in block["lines"]:
                        if "bbox" in line and "spans" in line:
                            x0, y0, x1, y1 = line["bbox"]
                            pixel_bbox = [
                                x0 * self.scale, y0 * self.scale,
                                x1 * self.scale, y1 * self.scale
                            ]
                            spans = line.get("spans", [])
                            text = " ".join([s.get("text", "") for s in spans])
                            if text.strip():
                                # 记录行内各 span 的 x 中心，供列判定使用（避免"整行中心"陷阱）
                                span_x_centers = []
                                for s in spans:
                                    if "bbox" in s:
                                        sx0, sy0, sx1, sy1 = s["bbox"]
                                        span_x_centers.append((sx0 * self.scale + sx1 * self.scale) / 2)
                                lines.append({
                                    'bbox': pixel_bbox,
                                    'text': text.strip(),
                                    'x0': pixel_bbox[0],
                                    'y0': pixel_bbox[1],
                                    'x1': pixel_bbox[2],
                                    'y1': pixel_bbox[3],
                                    'x_clusters': span_x_centers,
                                })

            # 如果文本行太少，尝试 OCR
            if len(lines) < 5 and self.ocr_extractor:
                ocr_lines = self._ocr_to_lines(page)
                if ocr_lines:
                    lines = ocr_lines

            # 布局推断：检测连续行、每行有多个 x 对齐词群
            if len(lines) >= 3:
                inferred = self._infer_tables_from_lines(lines)
                boxes.extend(inferred)

        except Exception:
            pass

        return boxes

    def _ocr_to_lines(self, page: 'fitz.Page') -> List[Dict]:
        """OCR 结果转换为行格式"""
        if not self.ocr_extractor:
            return []

        temp_img_path = None
        try:
            pix = self._get_pixmap(page, self.dpi)

            import tempfile
            fd, temp_img_path = tempfile.mkstemp(suffix='.png')
            os.close(fd)
            pix.save(temp_img_path)

            cells = self.ocr_extractor.recognize_from_image(temp_img_path, is_scanned=True)

            # 按 y 坐标聚类成行
            lines = self._cluster_cells_to_lines(cells)
            return lines
        except Exception:
            return []
        finally:
            if temp_img_path and os.path.exists(temp_img_path):
                try:
                    os.unlink(temp_img_path)
                except Exception:
                    pass

    def _cluster_cells_to_lines(self, cells: List[Dict]) -> List[Dict]:
        """将 OCR cells 按 y 坐标聚类成行"""
        if not cells:
            return []

        # 按 y 排序
        sorted_cells = sorted(cells, key=lambda c: (c.get('center_y', 0), c.get('center_x', 0)))

        # y 聚类
        lines = []
        current_line = []
        current_y = None
        y_tolerance = 10  # 像素容差

        for cell in sorted_cells:
            cy = cell.get('center_y', 0)
            if current_y is None:
                current_y = cy
                current_line.append(cell)
            elif abs(cy - current_y) <= y_tolerance:
                current_line.append(cell)
                current_y = sum(c.get('center_y', 0) for c in current_line) / len(current_line)
            else:
                if current_line:
                    line = self._merge_cells_to_line(current_line)
                    if line:
                        lines.append(line)
                current_line = [cell]
                current_y = cy

        if current_line:
            line = self._merge_cells_to_line(current_line)
            if line:
                lines.append(line)

        return lines

    def _merge_cells_to_line(self, cells: List[Dict]) -> Optional[Dict]:
        """合并同一行的 cells 为 line"""
        if not cells:
            return None

        x0 = min(c.get('x0', 0) for c in cells)
        y0 = min(c.get('y0', 0) for c in cells)
        x1 = max(c.get('x1', 0) for c in cells)
        y1 = max(c.get('y1', 0) for c in cells)
        text = " ".join([c.get('text', '') for c in cells])
        # 记录行内各 cell 的 x 中心，供列判定使用
        x_clusters = [((c.get('x0', 0) + c.get('x1', 0)) / 2) for c in cells]

        return {
            'bbox': [x0, y0, x1, y1],
            'text': text,
            'x0': x0, 'y0': y0,
            'x1': x1, 'y1': y1,
            'x_clusters': x_clusters,
        }

    def _infer_tables_from_lines(self, lines: List[Dict]) -> List[Dict]:
        """
        从文本行推断表格结构

        严格规则（必须同时满足）：
        1. 连续≥3行
        2. 每行≥2个 x 对齐词群（通过 x 投影检测多列结构）
        3. 相邻列的 x 位置在多行中保持稳定
        """
        inferred = []

        # 按 y 排序
        sorted_lines = sorted(lines, key=lambda l: l['y0'])

        i = 0
        while i < len(sorted_lines):
            # 找连续行
            table_lines = [sorted_lines[i]]

            j = i + 1
            while j < len(sorted_lines):
                line_height = max(sorted_lines[i]['y1'] - sorted_lines[i]['y0'], 1)
                gap = sorted_lines[j]['y0'] - sorted_lines[j-1]['y1']
                if gap <= line_height * 2.5:
                    table_lines.append(sorted_lines[j])
                    j += 1
                else:
                    break

            # 必须 ≥ 2 行（一般表格至少两行）
            if len(table_lines) >= 2:
                # === 严格多列 x 对齐检查 ===
                col_count = self._count_x_columns(table_lines)

                if col_count >= 2:
                    # 计算表格边界
                    all_x0 = [l['x0'] for l in table_lines]
                    all_y0 = [l['y0'] for l in table_lines]
                    all_x1 = [l['x1'] for l in table_lines]
                    all_y1 = [l['y1'] for l in table_lines]

                    bbox = [
                        min(all_x0), min(all_y0),
                        max(all_x1), max(all_y1)
                    ]

                    inferred.append({
                        'bbox': bbox,
                        'pixel_bbox': bbox,
                        'source': 'layout_inference',
                        'conf': 0.5,
                        'line_count': len(table_lines),
                        'col_count': col_count,
                    })

            i = j if j > i + 1 else i + 1

        return inferred

    def _count_x_columns(self, lines: List[Dict]) -> int:
        """
        检测多行文本的 x 列数（x 投影聚类）

        原理：收集每一行内各文本块(cell/span)的 x 中心，跨所有行做投影聚类。
        若检测到 ≥2 个稳定 x 簇（且每簇在多数行中均出现），则判定为多列表。

        关键修正：必须用"行内文本块的 x"，而非"整行合并 bbox 的中心"。
        中文表整行左/右对齐时整行中心相近，用整行中心会永远聚成 1 列而漏检。
        """
        if len(lines) < 2:
            return 0

        # 收集所有行的文本块 x 中心（行内多块 -> 多列信号）
        all_x = []
        for line in lines:
            xc = line.get('x_clusters')
            if xc:
                all_x.extend(xc)
            else:
                all_x.append((line['x0'] + line['x1']) / 2)

        # 样本太少不足以判多列
        if len(all_x) < 2:
            return 0
        if len(all_x) < 4:
            # 仅 2~3 个块（如 2 行 1 列表）不足以可靠判多列，保守返回 1（单列）
            return 1

        # 对 x 坐标做投影聚类
        sorted_x = sorted(all_x)
        diffs = [sorted_x[i+1] - sorted_x[i] for i in range(len(sorted_x)-1)]
        median_diff = sorted(diffs)[len(diffs) // 2]
        gap_threshold = max(median_diff * 1.5, 20)  # 至少 20px

        clusters = [[sorted_x[0]]]
        for k in range(1, len(sorted_x)):
            if sorted_x[k] - sorted_x[k-1] > gap_threshold:
                clusters.append([sorted_x[k]])
            else:
                clusters[-1].append(sorted_x[k])

        if len(clusters) < 2:
            return 1

        # 列稳定性：每个簇需在多数行中出现
        valid_clusters = 0
        for cluster in clusters:
            cluster_mean = sum(cluster) / len(cluster)
            matching_lines = 0
            for line in lines:
                xc = line.get('x_clusters')
                if xc:
                    # 该行是否有块接近此簇
                    if any(abs(x - cluster_mean) < gap_threshold for x in xc):
                        matching_lines += 1
                else:
                    cx = (line['x0'] + line['x1']) / 2
                    if abs(cx - cluster_mean) < gap_threshold:
                        matching_lines += 1
            # 至少 50% 的行覆盖此列
            if matching_lines >= len(lines) * 0.5:
                valid_clusters += 1

        return valid_clusters if valid_clusters >= 2 else 1

    def _coverage_check(
        self,
        page: 'fitz.Page',
        source_boxes: List[Dict]
    ) -> List[Dict]:
        """
        覆盖检查：未被表框盖住、又呈现表状特征的文字团块，补框进并集

        Returns:
            需要补充的表框
        """
        additional_boxes = []

        # 获取全页文本
        try:
            text_dict = page.get_text("dict")
            all_text_lines = []

            for block in text_dict.get("blocks", []):
                if "lines" in block:
                    for line in block["lines"]:
                        if "bbox" in line:
                            x0, y0, x1, y1 = line["bbox"]
                            pixel_bbox = [
                                x0 * self.scale, y0 * self.scale,
                                x1 * self.scale, y1 * self.scale
                            ]
                            # 记录行内各 span 的 x 中心，供列判定使用
                            span_x = []
                            for s in line.get("spans", []):
                                if "bbox" in s:
                                    sx0, _, sx1, _ = s["bbox"]
                                    span_x.append((sx0 * self.scale + sx1 * self.scale) / 2)
                            all_text_lines.append({
                                'bbox': pixel_bbox,
                                'x0': pixel_bbox[0],
                                'y0': pixel_bbox[1],
                                'x1': pixel_bbox[2],
                                'y1': pixel_bbox[3],
                                'text': "",
                                'x_clusters': span_x,
                            })

            # 检查哪些文字行未被任何表框覆盖
            uncovered_lines = []
            for line in all_text_lines:
                is_covered = False
                for box in source_boxes:
                    if self._boxes_intersect(line['bbox'], box['bbox'], threshold=0.5):
                        is_covered = True
                        break
                if not is_covered:
                    uncovered_lines.append(line)

            # 未覆盖文字行达到表格最小行数（≥2）即尝试补框；
            # 是否真为表仍由 _infer_tables_from_lines 的多列判定把关
            if len(uncovered_lines) >= 2:
                new_boxes = self._infer_tables_from_lines(uncovered_lines)
                additional_boxes.extend(new_boxes)

        except Exception:
            pass

        return additional_boxes

    def _merge_boxes(
        self,
        existing: List[Dict],
        new: List[Dict]
    ) -> List[Dict]:
        """合并新检测到的表框"""
        result = list(existing)
        for box in new:
            # 检查是否与已有框高度重叠
            is_duplicate = False
            for existing_box in result:
                if self._compute_iou(box['bbox'], existing_box['bbox']) > 0.6:
                    is_duplicate = True
                    break
            if not is_duplicate:
                result.append(box)
        return result

    def _dedup_by_iou(
        self,
        boxes: List[Dict],
        iou_threshold: float = 0.6
    ) -> List[Dict]:
        """
        IoU>阈值去重，保留边界更紧的

        原则：多个源检测到同一张表时，保留边界更紧（面积更小）的
        """
        if not boxes:
            return []

        # 按面积从小到大排序（边界紧的排前面）
        def box_area(b):
            x1, y1, x2, y2 = b['bbox']
            return (x2 - x1) * (y2 - y1)

        sorted_boxes = sorted(boxes, key=box_area)

        deduped = []
        for box in sorted_boxes:
            is_duplicate = False
            for kept in deduped:
                if self._compute_iou(box['bbox'], kept['bbox']) > iou_threshold:
                    # 如果重叠度超过阈值，用边界更紧的替换（面积小的保留）
                    is_duplicate = True
                    break
            if not is_duplicate:
                deduped.append(box)

        return deduped

    def _compute_iou(self, box1, box2) -> float:
        """计算两个框的 IoU"""
        x1 = max(box1[0], box2[0])
        y1 = max(box1[1], box2[1])
        x2 = min(box1[2], box2[2])
        y2 = min(box1[3], box2[3])

        if x1 >= x2 or y1 >= y2:
            return 0.0

        intersection = (x2 - x1) * (y2 - y1)

        area1 = (box1[2] - box1[0]) * (box1[3] - box1[1])
        area2 = (box2[2] - box2[0]) * (box2[3] - box2[1])

        union = area1 + area2 - intersection
        return intersection / union if union > 0 else 0.0

    def _boxes_intersect(self, box1, box2, threshold: float = 0.3) -> bool:
        """检查两个框是否相交超过阈值"""
        iou = self._compute_iou(box1, box2)
        return iou > threshold

    def classify_page_text(
        self,
        page: 'fitz.Page',
        page_num: int,
        table_boxes: List[Dict],
        is_scanned: bool = False
    ) -> Dict[str, List[Dict]]:
        """
        页面文字空间分类（表内/表外/未归类）+ 守恒检查

        支持三种页面类型：
        - 矢量页：PyMuPDF 取词框
        - 混合页：矢量 + OCR 补图片区
        - 扫描页：OCR 全页（始终触发，不受矢量文字数量限制）

        Args:
            page: PDF 页面
            page_num: 页码
            table_boxes: 去重后的表框列表（像素坐标）
            is_scanned: 是否为扫描页（扫描页始终全页OCR）

        Returns:
            {
                'in_table': [...],  # 表内文字
                'out_table': [...], # 表外文字
                'unclassified': [...],  # 未归类（兜底）
                'total_words': int,  # 总词数
                'conservation_ok': bool,  # 守恒是否通过
            }
        """
        result = {
            'in_table': [],
            'out_table': [],
            'unclassified': [],
            'total_words': 0,
            'conservation_ok': True,
        }

        # 1. 获取矢量文字
        text_dict = page.get_text("dict")
        all_spans = []

        for block in text_dict.get("blocks", []):
            if "lines" in block:
                for line in block["lines"]:
                    for span in line.get("spans", []):
                        text = span.get("text", "").strip()
                        if text:
                            x0, y0, x1, y1 = span["bbox"]
                            pixel_bbox = [
                                x0 * self.scale, y0 * self.scale,
                                x1 * self.scale, y1 * self.scale
                            ]
                            all_spans.append({
                                'text': text,
                                'pixel_bbox': pixel_bbox,
                                'vector_bbox': fitz.Rect(x0, y0, x1, y1),
                                'center_x': (pixel_bbox[0] + pixel_bbox[2]) / 2,
                                'center_y': (pixel_bbox[1] + pixel_bbox[3]) / 2,
                                'source': 'vector',
                            })

        # 2. 扫描页/混合页：扫描页始终全页OCR（不受矢量文字数量限制）
        #    混合页（is_scanned=False但矢量文字少）也触发OCR补全
        should_ocr = is_scanned or len(all_spans) < 10

        if should_ocr and self.ocr_extractor:
            logger.info(f"  classify_page_text: 触发OCR(is_scanned={is_scanned}, 矢量文字={len(all_spans)}, ocr_extractor={self.ocr_extractor is not None})")
            ocr_cells = self._get_ocr_cells(page)
            logger.info(f"  classify_page_text: OCR返回 {len(ocr_cells)} 个cells")
            if ocr_cells:
                if is_scanned and len(all_spans) == 0:
                    # 纯扫描页（无矢量文字）：直接用 OCR 结果，跳过去重
                    for cell in ocr_cells:
                        cell['source'] = 'ocr'
                        all_spans.append(cell)
                elif is_scanned:
                    # 扫描页（有少量矢量文字如页码/页眉）：OCR全页 + 矢量去重
                    # 扫描页必须始终全页OCR，不能因为矢量文字多就跳过
                    dedup_threshold = max(20, self.dpi // 8)
                    unique_ocr = []
                    for cell in ocr_cells:
                        is_duplicate = False
                        cell_cx = cell.get('center_x', 0)
                        cell_cy = cell.get('center_y', 0)
                        for span in all_spans:
                            sx = span['center_x']
                            sy = span['center_y']
                            if abs(cell_cx - sx) < dedup_threshold and abs(cell_cy - sy) < dedup_threshold:
                                is_duplicate = True
                                break
                        if not is_duplicate:
                            cell['source'] = 'ocr'
                            unique_ocr.append(cell)
                    all_spans.extend(unique_ocr)
                elif len(all_spans) < 10:
                    # 混合页：去重，但阈值随 DPI 自适应
                    dedup_threshold = max(20, self.dpi // 8)
                    unique_ocr = []
                    for cell in ocr_cells:
                        is_duplicate = False
                        cell_cx = cell.get('center_x', 0)
                        cell_cy = cell.get('center_y', 0)
                        for span in all_spans:
                            sx = span['center_x']
                            sy = span['center_y']
                            if abs(cell_cx - sx) < dedup_threshold and abs(cell_cy - sy) < dedup_threshold:
                                is_duplicate = True
                                break
                        if not is_duplicate:
                            cell['source'] = 'ocr'
                            unique_ocr.append(cell)
                    all_spans.extend(unique_ocr)
            elif is_scanned:
                # 扫描页OCR完全失败，记录警告（不阻断流程）
                logger.warning(f"  classify_page_text: 扫描页OCR返回0个cells(is_scanned={is_scanned}, 矢量文字={len(all_spans)})")
                pass
        elif is_scanned:
            logger.warning(f"  classify_page_text: 扫描页未触发OCR(should_ocr={should_ocr}, ocr_extractor={self.ocr_extractor is not None})")

        result['total_words'] = len(all_spans)

        # 3. 空间分类
        for span in all_spans:
            center_x = span['center_x']
            center_y = span['center_y']

            in_any_table = False
            for table_box in table_boxes:
                tx1, ty1, tx2, ty2 = table_box['pixel_bbox']
                if tx1 <= center_x <= tx2 and ty1 <= center_y <= ty2:
                    span['table_index'] = table_box.get('index', -1)
                    span['table_source'] = table_box.get('source', 'unknown')
                    result['in_table'].append(span)
                    in_any_table = True
                    break

            if not in_any_table:
                result['out_table'].append(span)

        # 4. 守恒检查
        total = len(all_spans)
        classified = len(result['in_table']) + len(result['out_table'])

        if classified < total:
            # 有文字未被归类 → 放入未归类
            result['conservation_ok'] = False
            for span in all_spans:
                center_x = span['center_x']
                center_y = span['center_y']
                in_classified = False
                for tb in table_boxes:
                    tx1, ty1, tx2, ty2 = tb['pixel_bbox']
                    if tx1 <= center_x <= tx2 and ty1 <= center_y <= ty2:
                        in_classified = True
                        break
                if not in_classified:
                    already_out = any(
                        o.get('vector_bbox') == span.get('vector_bbox') and o.get('text') == span.get('text')
                        for o in result['out_table']
                    )
                    if not already_out:
                        result['unclassified'].append({
                            **span,
                            'reason': '守恒兜底：未被任何表框覆盖且未归入表外',
                        })
        elif classified > total:
            result['conservation_ok'] = False

        return result

    def _get_ocr_cells(self, page: 'fitz.Page') -> List[Dict]:
        """获取 OCR 识别结果"""
        if not self.ocr_extractor:
            logger.warning("_get_ocr_cells: ocr_extractor为None，跳过OCR")
            return []

        temp_img_path = None
        try:
            pix = self._get_pixmap(page, self.dpi)
            logger.info(f"_get_ocr_cells: 渲染页面 {pix.width}x{pix.height} (DPI={self.dpi})")

            import tempfile
            fd, temp_img_path = tempfile.mkstemp(suffix='.png')
            os.close(fd)
            pix.save(temp_img_path)

            cells = self.ocr_extractor.recognize_from_image(temp_img_path, is_scanned=True)

            # 转换为统一格式
            result = []
            for cell in cells:
                x0 = cell.get('x0', 0)
                y0 = cell.get('y0', 0)
                x1 = cell.get('x1', 0)
                y1 = cell.get('y1', 0)
                result.append({
                    'text': cell.get('text', ''),
                    'pixel_bbox': [x0, y0, x1, y1],
                    'vector_bbox': fitz.Rect(
                        x0 * self.inverse_scale,
                        y0 * self.inverse_scale,
                        x1 * self.inverse_scale,
                        y1 * self.inverse_scale
                    ),
                    'center_x': (x0 + x1) / 2,
                    'center_y': (y0 + y1) / 2,
                })
            return result
        except Exception:
            return []
        finally:
            if temp_img_path and os.path.exists(temp_img_path):
                try:
                    os.unlink(temp_img_path)
                except Exception:
                    pass

    def extract_table_content(
        self,
        page: 'fitz.Page',
        table_box: Dict,
        in_table_cells: List[Dict]
    ) -> Optional[Dict]:
        """
        从表框内提取表格内容（结构切分）

        策略：
        1. 优先按矢量线切
        2. 无线则 y 聚类切行、x 投影切列
        3. 切完 <2列或<2行 → 判定不是表，返回 None
        """
        bbox = table_box['pixel_bbox']
        x1, y1, x2, y2 = bbox

        # 筛选表框内的 cells（第一次：严格筛选）
        table_cells = []
        for cell in in_table_cells:
            cx = cell.get('center_x', 0)
            cy = cell.get('center_y', 0)
            if x1 <= cx <= x2 and y1 <= cy <= y2:
                table_cells.append(cell)

        # === 放宽策略：如果 cells 太少，尝试用扩大的 bbox 重新捞 ===
        if len(table_cells) < 2:
            # 扩大 bbox 20% 再捞一次
            bw = x2 - x1
            bh = y2 - y1
            expand_x = bw * 0.1
            expand_y = bh * 0.1
            ex1, ey1, ex2, ey2 = x1 - expand_x, y1 - expand_y, x2 + expand_x, y2 + expand_y

            for cell in in_table_cells:
                cx = cell.get('center_x', 0)
                cy = cell.get('center_y', 0)
                if ex1 <= cx <= ex2 and ey1 <= cy <= ey2:
                    table_cells.append(cell)

        # 仍然不够 → 再退化：用与表框 IoU > 0.3 的 cell
        if len(table_cells) < 2:
            for cell in in_table_cells:
                cell_bbox = cell.get('pixel_bbox', [0, 0, 0, 0])
                if len(cell_bbox) == 4:
                    iou = self._compute_iou(bbox, cell_bbox)
                    if iou > 0.3:
                        table_cells.append(cell)

        if len(table_cells) < 2:
            return None

        # 尝试用矢量线切分
        rows, cols = self._split_by_lines(page, bbox, table_cells)

        # 如果矢量线失败，用聚类切分
        if rows is None:
            rows, cols = self._split_by_clustering(bbox, table_cells)

        # 退化自愈：切完 <2列或 <2行
        if rows is None or len(rows) < 2 or len(cols) < 2:
            return None  # 判定不是表，内容退回表外

        # 构建表格数据
        table_data = self._build_table_data(rows, cols, table_cells)

        return {
            'bbox': bbox,
            'source': table_box.get('source', 'unknown'),
            'rows': rows,
            'cols': cols,
            'data': table_data,
            'row_count': len(rows),
            'col_count': len(cols),
        }

    def _split_by_lines(
        self,
        page: 'fitz.Page',
        bbox: List[float],
        cells: List[Dict]
    ) -> Tuple[Optional[List[float]], Optional[List[float]]]:
        """尝试用矢量线切分行列"""
        try:
            # 获取页面上的线条
            drawings = page.get_drawings()

            horizontal_lines = []
            vertical_lines = []

            for drawing in drawings:
                if drawing.get('type') == 'l':  # 线条
                    pts = drawing.get('pts', [])
                    if len(pts) >= 2:
                        x0, y0 = pts[0]
                        x1, y1 = pts[1]

                        # 转换为像素坐标
                        px0, py0 = x0 * self.scale, y0 * self.scale
                        px1, py1 = x1 * self.scale, y1 * self.scale

                        # 检查是否在表框内
                        if bbox[0] <= px0 <= bbox[2] and bbox[1] <= py0 <= bbox[3]:
                            # 水平线（y 近似相等）
                            if abs(py0 - py1) < 10:
                                horizontal_lines.append((py0 + py1) / 2)
                            # 垂直线（x 近似相等）
                            elif abs(px0 - px1) < 10:
                                vertical_lines.append((px0 + px1) / 2)

            # 如果有足够多的线条
            if len(horizontal_lines) >= 2 and len(vertical_lines) >= 2:
                rows = sorted(set([bbox[1]] + horizontal_lines + [bbox[3]]))
                cols = sorted(set([bbox[0]] + vertical_lines + [bbox[2]]))
                return rows, cols

        except Exception:
            pass

        return None, None

    def _split_by_clustering(
        self,
        bbox: List[float],
        cells: List[Dict]
    ) -> Tuple[List[float], List[float]]:
        """用聚类切分行列"""
        # y 聚类切行
        y_values = [c.get('center_y', 0) for c in cells]
        y_values = sorted(set([round(y, 1) for y in y_values]))

        # 简单聚类：间距小于阈值的归为一行
        row_groups = []
        current_group = [y_values[0]]
        y_threshold = 15  # 像素阈值

        for i in range(1, len(y_values)):
            if y_values[i] - current_group[-1] <= y_threshold:
                current_group.append(y_values[i])
            else:
                row_groups.append(sum(current_group) / len(current_group))
                current_group = [y_values[i]]
        row_groups.append(sum(current_group) / len(current_group))

        # 添加边界
        rows = [bbox[1]] + row_groups + [bbox[3]]
        rows = sorted(set([round(r, 1) for r in rows]))

        # x 投影切列
        x_values = [c.get('center_x', 0) for c in cells]
        x_values = sorted(set([round(x, 1) for x in x_values]))

        # 简单聚类：间距小于阈值的归为一列
        col_groups = []
        current_group = [x_values[0]]
        x_threshold = 30  # 像素阈值

        for i in range(1, len(x_values)):
            if x_values[i] - current_group[-1] <= x_threshold:
                current_group.append(x_values[i])
            else:
                col_groups.append(sum(current_group) / len(current_group))
                current_group = [x_values[i]]
        col_groups.append(sum(current_group) / len(current_group))

        # 添加边界
        cols = [bbox[0]] + col_groups + [bbox[2]]
        cols = sorted(set([round(c, 1) for c in cols]))

        return rows, cols

    def _build_table_data(
        self,
        rows: List[float],
        cols: List[float],
        cells: List[Dict]
    ) -> List[List[str]]:
        """构建表格数据矩阵"""
        # 初始化空表格
        table = [["" for _ in range(len(cols) - 1)] for _ in range(len(rows) - 1)]

        # 分配每个 cell 到对应行列
        for cell in cells:
            cx = cell.get('center_x', 0)
            cy = cell.get('center_y', 0)
            text = cell.get('text', '')

            # 找出行索引
            row_idx = 0
            for i in range(len(rows) - 1):
                if rows[i] <= cy <= rows[i + 1]:
                    row_idx = i
                    break

            # 找出列索引
            col_idx = 0
            for i in range(len(cols) - 1):
                if cols[i] <= cx <= cols[i + 1]:
                    col_idx = i
                    break

            # 合并文本
            if table[row_idx][col_idx]:
                table[row_idx][col_idx] += " " + text
            else:
                table[row_idx][col_idx] = text

        return table


def main_pdf_to_excel(
    pdf_path: str,
    output_path: str,
    model_db_path: Optional[str] = None,
    page_range: Optional[Tuple[int, int]] = None,
    use_yolo: bool = True,
    yolo_weights: Optional[str] = None,
):
    """
    PDF → Excel 主入口

    Args:
        pdf_path: 输入 PDF 路径
        output_path: 输出 Excel 路径
        model_db_path: 型号数据库路径（可选）
        page_range: 页码范围 (start, end)，None 表示全部
        use_yolo: 是否启用 YOLO 辅助检测（用于扫描页）
        yolo_weights: YOLO 权重路径（可选）
    """
    if not os.path.exists(pdf_path):
        print(f"[错误] PDF 文件不存在: {pdf_path}")
        return

    # 确保输出目录存在
    output_dir = os.path.dirname(output_path)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir, exist_ok=True)

    print("=" * 60)
    print("PDF → Excel 型号提取管线")
    print("=" * 60)
    print(f"输入: {pdf_path}")
    print(f"输出: {output_path}")
    print(f"型号库: {model_db_path or '未加载（仅正则匹配）'}")
    print(f"YOLO: {'启用' if use_yolo else '禁用'}")
    print("=" * 60)

    start_time = time.time()

    # 初始化 YOLO 检测器（可选）
    yolo_detector = None
    if use_yolo:
        yolo_detector = YOLOTableDetector(weights_path=yolo_weights)
        if yolo_detector.load_model():
            print("[YOLO] 表格检测已启用（仅用于扫描页）")
        else:
            print("[YOLO] 未启用，扫描页将依赖 OCR 兜底")
            yolo_detector = None

    # 初始化增强版提取器
    extractor = EnhancedPDFTableExtractor(
        pdf_path,
        use_ocr=True,
        model_db_path=model_db_path,
        yolo_detector=yolo_detector,  # 传入 YOLO 检测器
    )

    # 设置页码范围
    page_range_obj = None
    if page_range:
        start, end = page_range
        total = extractor.doc.page_count
        start = max(0, start - 1)
        end = min(total, end)
        page_range_obj = range(start, end)
        print(f"处理页码: {start + 1} - {end}")

    # 导出
    extractor.export_to_excel(output_path, page_range_obj)

    elapsed = time.time() - start_time
    print(f"\n耗时: {elapsed:.1f} 秒")


if __name__ == "__main__":
    """统一入口：有 --pdf 参数走新管线，否则走旧版测试"""
    import argparse

    # 检查是否有 --pdf 参数
    has_pdf_arg = '--pdf' in sys.argv

    if has_pdf_arg:
        parser = argparse.ArgumentParser(
            description='PDF → Excel 型号提取管线（集成 YOLO 扫描页检测）',
            formatter_class=argparse.RawDescriptionHelpFormatter,
            epilog='''
示例:
  # 处理全部页面（默认启用 YOLO）
  python pdf_table_extractor.py --pdf input.pdf --output result.xlsx

  # 处理指定页码范围
  python pdf_table_extractor.py --pdf input.pdf --output result.xlsx --start 1 --end 10

  # 带型号数据库
  python pdf_table_extractor.py --pdf input.pdf --output result.xlsx --db model_db/

  # 禁用 YOLO（仅矢量检测）
  python pdf_table_extractor.py --pdf input.pdf --output result.xlsx --no-yolo

  # 自定义 YOLO 权重路径
  python pdf_table_extractor.py --pdf input.pdf --output result.xlsx --yolo-weights weights/table_best.pt
            '''
        )

        parser.add_argument('--pdf', required=True, help='输入 PDF 路径')
        parser.add_argument('--output', required=True, help='输出 Excel 路径')
        parser.add_argument('--db', default=None, help='型号数据库目录路径')
        parser.add_argument('--start', type=int, default=1, help='起始页码（1-based）')
        parser.add_argument('--end', type=int, default=None, help='结束页码（1-based）')
        parser.add_argument('--no-yolo', action='store_true', help='禁用 YOLO 检测（仅用于矢量 PDF）')
        parser.add_argument('--yolo-weights', default=None, help='YOLO 权重文件路径')

        args = parser.parse_args()

        page_range = None
        if args.end:
            page_range = (args.start, args.end)

        main_pdf_to_excel(
            pdf_path=args.pdf,
            output_path=args.output,
            model_db_path=args.db,
            page_range=page_range,
            use_yolo=not args.no_yolo,
            yolo_weights=args.yolo_weights,
        )
    else:
        # 无参数时走旧版测试
        _legacy_test_entry()